# -*- coding=utf-8 -*-

from django.db.models import Q
from django.db import transaction

from models import *
from user_center.utils.constant import *
from user_center.utils.err_code import *
from user_center.utils.public_fun import *
from user_center.utils.request_auth import *
from user_center.apps.account.agents import *
from user_center.apps.school.models import *
from user_center.utils.file_fun import *
from user_center.apps.common.agents import check_verify_code
import openpyxl
from openpyxl.styles import Font


def list_student_parent(user):
    school_id = user.school_id
    student_obj = Student.objects.filter(school_id=school_id, account_id=user.id, del_flag=NO).first()
    if not student_obj or not student_obj.account:
        return {"c": ERR_USER_AUTH[0], "m": ERR_USER_AUTH[1], "d": []}
    return list_parent(user=user, student_id=student_obj.id, verbose="1")


def list_parent(user, full_name="", student_name="", student_id="", verbose="", grade_name="", class_id="",
                is_active="", display_mode=PARENT_LIST_ONE_PARENT_PER_LINE):
    school_id = user.school_id

    # 检查用户角色班主任只能管理本班级学生
    if ADMIN_USER_TYPE_CLASS_ADMIN in user.role:
        dict_resp = check_cls_permission(user, -1)
        if dict_resp["c"] != ERR_SUCCESS[0]:
            return dict_resp
        else:
            class_id = dict_resp["d"][0]

    parent_list = Parent.objects.filter(school_id=school_id, del_flag=NO).all()
    parent_student_list = ParentStudent.objects.filter(status=APPLICATION_STATUS_APPROVED[0], student__del_flag=NO, del_flag=NO).all()
    # 清除不存在关联学生的家长信息
    # all_parent_id_list = parent_student_list.values_list("parent_id", flat=True)
    # Parent.objects.exclude(id__in=all_parent_id_list).delete()
    if not verbose or verbose == "0":  # 默认不显示已经毕业学生家长信息
        parent_student_list = parent_student_list.filter(student__cls__graduate_status=NO)
    if is_active and is_active == "0":
        parent_student_list = parent_student_list.filter(parent__is_active=NO)
    else:
        parent_student_list = parent_student_list.filter(parent__is_active=YES)
    if full_name:
        parent_list = parent_list.filter(full_name__contains=full_name)
    if student_name:
        parent_student_list = parent_student_list.filter(student__full_name__contains=student_name)
    if grade_name:
        parent_student_list = parent_student_list.filter(student__cls__grade_name=grade_name)
    if class_id:
        parent_student_list = parent_student_list.filter(student__cls_id=class_id)
    if student_id:
        parent_student_list = parent_student_list.filter(student_id=int(student_id))
    parent_id_list = parent_student_list.values_list("parent_id", flat=True)
    parent_list = parent_list.filter(id__in=parent_id_list)

    parent_list = parent_list.values('id', 'account_id', 'account__username', 'full_name', 'id_card',
                                     'sex', 'account__mobile', 'email', 'address',
                                     'company', 'comments', "image__url", 'account__last_login')
    parent_id_list = map(lambda x: x["id"], parent_list)
    student_list = ParentStudent.objects.filter(parent_id__in=parent_id_list, del_flag=NO).\
        values("parent_id", "student_id", "student__full_name", "student__account__code", "student__cls__class_name",
               "student__cls_id", "student__cls__grade_name", "relation", "comments")
    student_dict = {}
    for student in student_list:
        student_info = dict(student_id=str(student["student_id"]), student_full_name=student["student__full_name"],
                            student_code=student["student__account__code"], relation=student["relation"],
                            comments=student["comments"], student_class_name=student["student__cls__class_name"],
                            class_id=student["student__cls_id"], grade_name=student["student__cls__grade_name"])
        if student["parent_id"] in student_dict.keys():
            student_dict[student["parent_id"]].append(student_info)
        else:
            student_dict[student["parent_id"]] = [student_info]
    ret_parent_list = []
    if display_mode == PARENT_LIST_ONE_PARENT_PER_LINE:
        ret_parent_list = []
        for parent in parent_list:
            ret_parent_info = dict(id=str(parent["id"]),
                                   account_id=str(parent["account_id"]),
                                   username=parent['account__username'],
                                   id_card=parent['id_card'],
                                   full_name=parent['full_name'],
                                   sex=parent['sex'],
                                   email=parent['email'],
                                   mobile=parent['account__mobile'],
                                   address=parent['address'],
                                   company=parent['company'],
                                   image_url="",
                                   last_login=datetime_to_str(parent['account__last_login'], DATE_FORMAT_TIME),
                                   children=student_dict[parent["id"]])
            if parent["image__url"]:
                ret_parent_info["image_url"] = get_image_url(parent["image__url"])
            ret_parent_list.append(ret_parent_info)
    else:
        for parent in parent_list:
            parent_info = dict(id=parent["id"],
                               account_id=parent["account_id"],
                               username=parent['account__username'],
                               id_card=parent['id_card'],
                               full_name=parent['full_name'],
                               sex=parent['sex'],
                               email=parent['email'],
                               mobile=parent['account__mobile'],
                               address=parent['address'],
                               company=parent['company'],
                               image_url="",
                               last_login=datetime_to_str(parent['account__last_login'], DATE_FORMAT_TIME))
            if parent["image__url"]:
                ret_parent_info["image_url"] = get_image_url(parent["image__url"])
            for student in student_dict[parent["id"]]:
                ret_parent_info = parent_info.copy()
                ret_parent_info["student_id"] = student["student_id"]
                ret_parent_info["student_full_name"] = student["student_full_name"]
                ret_parent_info["student_code"] = student["student_code"]
                ret_parent_info["relation"] = student["relation"]
                ret_parent_info["comments"] = student["comments"]
                ret_parent_list.append(ret_parent_info)

    dict_resp = {"c": ERR_SUCCESS[0], "m": ERR_SUCCESS[1], "d": ret_parent_list}
    return dict_resp


def detail_parent(user, parent_id="", account_id=""):
    school_id = user.school_id

    parent = Parent.objects.filter(school_id=school_id, id=int(parent_id), del_flag=NO).\
        values('id', 'account_id', 'account__username', 'full_name', 'id_card',
               'sex', 'account__mobile', 'email', 'address',
               'company', 'comments', 'image__url', "birthday", "native_place").first()

    # 检查自己修改自己的信息
    if not parent:
        return {"c": ERR_USER_NOT_EXIST[0], "m": ERR_USER_NOT_EXIST[1], "d": []}

    if account_id and int(account_id) != parent["account_id"]:
        return {"c": ERR_USER_AUTH[0], "m": ERR_USER_AUTH[1], "d": []}

    student_list = ParentStudent.objects.filter(parent_id=parent_id, del_flag=NO).\
        values("parent_id", "student_id", "student__full_name", "student__account__code", "relation", "comments")
    ret_student_list = []
    for student in student_list:
        student_info = dict(student_id=str(student["student_id"]), student_full_name=student["student__full_name"],
                            student_code=student["student__account__code"], relation=student["relation"],
                            comments=student["comments"])
        ret_student_list.append(student_info)

    ret_parent_info = dict(id=str(parent["id"]),
                           account_id=str(parent["account_id"]),
                           username=parent['account__username'],
                           id_card=parent['id_card'],
                           full_name=parent['full_name'],
                           sex=parent['sex'],
                           email=parent['email'],
                           mobile=parent['account__mobile'],
                           address=parent['address'],
                           company=parent['company'],
                           birthday=datetime_to_str(parent["birthday"]),
                           native_place=parent["native_place"],
                           image_url="",
                           children=ret_student_list
                           )
    if parent["image__url"]:
        ret_parent_info["image_url"] = get_image_url(parent["image__url"])
    return {"c": ERR_SUCCESS[0], "m": ERR_SUCCESS[1], "d": [ret_parent_info]}


@transaction.atomic
def add_parent(user, parent_info, is_override=False):
    school_id = user.school_id

    # username = parent_info.get("username", "")
    password = parent_info.get("password", "")
    full_name = clean_string(parent_info.get("full_name", ""))
    id_card = parent_info.get("id_card", "")
    sex = parent_info.get("sex", "")
    mobile = parent_info.get("mobile", "")
    email = parent_info.get("email", "")
    address = parent_info.get("address", "")
    company = parent_info.get("company", "")
    comments = parent_info.get("comments", "")
    birthday = parent_info.get("birthday", "")
    native_place = parent_info.get("native_place", "")

    image_url = parent_info.get("image_url", "")
    image_obj = None
    if image_url:
        image_url = get_file_name_from_url(image_url)
        image_obj = Image.objects.filter(url__contains=image_url, del_flag=NO).first()
        if not image_obj:
            return {"c": ERR_USER_INFO_IMAGE_NOT_EXIST[0], "m": ERR_USER_INFO_IMAGE_NOT_EXIST[1], "d": []}

    # 检查必填项
    if not mobile or not full_name:
        logger.error(" parent info is not integrity")
        return {"c": ERR_USER_INFO_INCOMPLETE[0], "m": ERR_USER_INFO_INCOMPLETE[1], "d": []}

    # 检查用户角色班主任只能管理本班级学生
    cls_id = None
    if ADMIN_USER_TYPE_CLASS_ADMIN in user.role and len(user.role) == 1:
        teacher_obj = Teacher.objects.filter(school_id=user.school_id, account_id=user.id, del_flag=NO).first()
        if teacher_obj:
            cls_id = teacher_obj.cls_id

    # 检查帐号在同一个学校的同一个角色是否已经存在
    parent = Parent.objects.filter(school_id=school_id, account__mobile=mobile, del_flag=NO).first()
    if not parent:
        # 创建帐号或获取已存在的帐号
        err_code = add_account(mobile=mobile, password=password, type=USER_TYPE_PARENT, school_id=school_id)
        if err_code[0] != ERR_SUCCESS[0]:
            return {"c": err_code[0], "m": err_code[1], "d": []}
        account_id = err_code[1]

        # 添加家长到Parent表
        parent = Parent.objects.create(school_id=school_id, account_id=account_id)
        if full_name: parent.full_name = full_name
        if id_card: parent.id_card = id_card
        if email: parent.email = email
        if sex: parent.sex = sex
        if address: parent.address = address
        if image_obj: parent.image = image_obj
        if company: parent.company = company
        if comments: parent.comments = comments
        if birthday: parent.birthday = str_to_datetime(birthday)
        if native_place: parent.native_place = native_place
        if image_obj: parent.image = image_obj
        parent.save()

    # 检查身份证号是否已经占用
    if id_card and check_id_code_have_occupied(id_card, parent.account_id):
        raise Exception(ERR_USER_ID_CARD_CONFLICT_ERROR[1])

    # 添加学生和家长的对应关系
    for child in parent_info["children"]:
        student_id = child.get("student_id", "")
        student_code = child.get("student_code", "")
        relation = child.get("relation", "")
        comments = child.get("comments", "")
        status = child.get("status", "")

        # 检查并获取学生对象
        student_obj = None
        if student_id:
            student_obj = Student.objects.filter(school_id=school_id, id=int(student_id), del_flag=NO).first()
        elif student_code:
            student_obj = Student.objects.filter(school_id=school_id, account__code=student_code, del_flag=NO).first()
        else:
            logger.error(" parent children info is not integrity")
            raise Exception(ERR_USER_INFO_INCOMPLETE[1])
        if not student_obj:
            logger.error(" parent children info is not integrity")
            raise Exception(ERR_USER_INFO_INCOMPLETE[1])
        if cls_id and student_obj.cls_id != cls_id:
            logger.error("class master cannot add other class student parent")
            raise Exception(ERR_USER_AUTH[1])

        # add to ParentStudent table
        parent_student, created = ParentStudent.objects.update_or_create(parent_id=parent.id, student_id=student_obj.id, del_flag=NO)
        # if not parent_student:
        #     parent_student = ParentStudent.objects.create(parent_id=parent.id, student_id=student_obj.id)
        if relation: parent_student.relation = relation
        if comments: parent_student.comments = comments
        if status: parent_student.status = status
        parent_student.save()

    return {"c": ERR_SUCCESS[0], "m": ERR_SUCCESS[1], "d": []}


@transaction.atomic
def delete_parent(user, parent_id_list):
    school_id = user.school_id

    # id_list and account_id_list should be delete
    id_list = []
    account_id_list = []
    parent_id_list = map(lambda x: int(x), parent_id_list)
    parent_list = Parent.objects.filter(id__in=parent_id_list, school_id=school_id,
                                        del_flag=NO).values("id", "account_id")
    for parent in parent_list:
        id_list.append(parent["id"])
        account_id_list.append(parent["account_id"])
    # delete parent in parent table
    Parent.objects.filter(id__in=id_list, del_flag=NO).update(del_flag=YES, update_time=datetime.datetime.now())
    ParentStudent.objects.filter(parent_id__in=id_list, del_flag=NO).update(del_flag=YES, update_time=datetime.datetime.now())

    # delete Account if necessary
    for account_id in account_id_list:
        check_and_delete_account(account_id)

    # check if delete all
    err_num = len(parent_id_list) - len(id_list)
    if err_num > 0:
        msg = u"%d条记录删除失败，可能是权限不够" % err_num
        dict_resp = {"c": -1, "m": msg, "d": []}
        return dict_resp

    return {"c": ERR_SUCCESS[0], "m": ERR_SUCCESS[1], "d": []}


@transaction.atomic
def active_parent(user, parent_id_list, is_active=""):
    school_id = user.school_id
    parent_id_list = map(lambda x: int(x), parent_id_list)

    total_num = len(parent_id_list)
    real_num = 0
    for parent_id in parent_id_list:
        parent_obj = Parent.objects.filter(school_id=school_id, id=parent_id, del_flag=FLAG_NO).first()
        if not parent_obj:
            continue
        if is_active and int(is_active) == 0:
            parent_obj.is_active = FLAG_NO
        else:
            parent_obj.is_active = FLAG_YES
        parent_obj.save()
        real_num += 1

    # return msg
    if real_num == total_num:
        return {"c": ERR_SUCCESS[0], "m": ERR_SUCCESS[1], "d": []}
    elif total_num == 1:
        return {"c": ERR_OP_ERROR[0], "m": ERR_OP_ERROR[1], "d": []}
    else:
        msg = u"%d条记录操作失败，%d条记录操作成功" % (total_num-real_num, real_num)
        return {"c": ERR_OP_PART_ERROR[0], "m": msg, "d": []}


@transaction.atomic
def update_parent(user, parent_info, account_id=""):
    school_id = user.school_id

    parent_id = parent_info.get("id", "")
    if not parent_id:
        return {"c": ERR_USER_NOT_EXIST[0], "m": ERR_USER_NOT_EXIST[1], "d": []}
    parent = Parent.objects.filter(id=int(parent_id), school_id=school_id, del_flag=NO).first()
    if not parent:
        return {"c": ERR_USER_NOT_EXIST[0], "m": ERR_USER_NOT_EXIST[1], "d": []}

    # 若account_id为自己修改自己的信息
    if account_id and int(account_id) != parent.account_id:
        return {"c": ERR_USER_AUTH[0], "m": ERR_USER_AUTH[1], "d": []}

    username = parent_info.get("username")
    full_name = clean_string(parent_info.get("full_name"))
    id_card = parent_info.get("id_card")
    sex = parent_info.get("sex")
    mobile = parent_info.get("mobile")
    email = parent_info.get("email")
    address = parent_info.get("address")
    company = parent_info.get("company")
    comments = parent_info.get("comments")
    birthday = parent_info.get("birthday", "")
    native_place = parent_info.get("native_place", "")
    image_url = parent_info.get("image_url")

    image_obj = None
    if image_url:
        image_url = get_file_name_from_url(image_url)
        image_obj = Image.objects.filter(url__contains=image_url, del_flag=NO).first()
        if not image_obj:
            return {"c": ERR_USER_INFO_IMAGE_NOT_EXIST[0], "m": ERR_USER_INFO_IMAGE_NOT_EXIST[1], "d": []}

    # 检查身份证号是否已经占用
    if id_card and check_id_code_have_occupied(id_card, parent.account_id):
        raise Exception(ERR_USER_ID_CARD_CONFLICT_ERROR[1])

    # update account
    account = parent.account
    err_code = update_account(user=user, account=account, username=username, mobile=mobile)
    if err_code != ERR_SUCCESS:
        return {"c": err_code[0], "m": err_code[1], "d": []}

    # 更新Parent表
    if full_name: parent.full_name = full_name
    if id_card: parent.id_card = id_card
    elif id_card == "": parent.id_card = id_card
    if email: parent.email = email
    elif email == "": parent.email = email
    if sex: parent.sex = sex
    if address: parent.address = address
    elif address == "": parent.address = address
    if company: parent.company = company
    elif company == "": parent.company = company
    if comments: parent.comments = comments
    elif comments == "": parent.comments = comments
    if birthday: parent.birthday = str_to_datetime(birthday)
    if native_place: parent.native_place = native_place
    if image_obj: parent.image = image_obj
    parent.save()

    # 更新学生和家长的对应关系
    children = parent_info.get("children", [])
    if not children:
        return {"c": ERR_SUCCESS[0], "m": ERR_SUCCESS[1], "d": []}
    org_id_list = list(ParentStudent.objects.filter(parent_id=parent.id, del_flag=NO).values_list("student_id", flat=True))
    for child in children:
        student_id = int(child.get("student_id", ""))
        relation = child.get("relation", "")
        comments = child.get("comments", "")
        # 忽略不存在的学生
        if not student_id or not Student.objects.filter(id=student_id, del_flag=NO).exists():
            continue
        if student_id in org_id_list:
            org_id_list.remove(student_id)
        # 创建新的学生家长关系
        student = ParentStudent.objects.filter(parent_id=parent.id, student_id=student_id, del_flag=NO).first()
        if not student:
            student = ParentStudent.objects.create(parent_id=parent.id, student_id=student_id)
        student.relation = relation
        student.comments = comments
        student.save()
    # 删除之前不用的家长学生关系
    ParentStudent.objects.filter(parent_id=parent.id, student_id__in=org_id_list, del_flag=NO).\
        update(del_flag=YES, update_time=datetime.datetime.now())
    return {"c": ERR_SUCCESS[0], "m": ERR_SUCCESS[1], "d": []}


def import_parent(user, file_obj, is_override=False):
    school_id = user.school_id
    ret_data = []
    skip_id_list = []
    success_num = 0

    if not file_obj.name.endswith('xlsx'):
        return {"c": ERR_FILE_FORMAT_NOT_SUPPORTED[0], "m": ERR_FILE_FORMAT_NOT_SUPPORTED[1],  "d": []}

    # write file data to tmp path
    file_path = gen_path()
    with open(file_path, 'wb+') as destination:
        for chunk in file_obj.chunks():
            destination.write(chunk)

    wb = openpyxl.load_workbook(file_path)
    ws = wb.get_active_sheet()

    property_display_list = [u"姓名*", u"性别*", u"手机号*", u"学生学籍号*", u"学生姓名",
                             u"所属关系", u"身份证号"]
    property_name_list = ["full_name", "sex", "mobile", "student_code", "student_ful_name",
                          "relation",  "id_card"]
    property_len = len(property_name_list)

    # check the first row
    for i in range(property_len):
        if property_display_list[i] != ws.cell(column=i+1, row=1).value:
            return {"c": ERR_FILE_TEMPLATE_ERROR[0], "m": ERR_FILE_TEMPLATE_ERROR[1],  "d": []}
    row_num = 0
    if len(ws.rows) > MAX_IMPORT_PRENT_NUM:
        err_msg = ERR_IMPORT_NUM_ERROR[1] + ": "
        err_msg += u"最多一次导入%d条记录" % MAX_IMPORT_PRENT_NUM
        return {"c": ERR_IMPORT_NUM_ERROR[0], "m": err_msg, "d": []}
    for row in ws.rows:
        row_num += 1
        if row_num == 1:
            # 第一行为标题栏
            continue
        col_num = 0
        parent_info = {"children":[{}]}
        for cell in row:
            if col_num >= property_len:
                break
            property_name = property_name_list[col_num]
            value = ""
            if isinstance(cell.value, (int, long)):
                value = str(cell.value)
            elif isinstance(cell.value, basestring):
                value = cell.value.strip()
            if property_name in ["student_code", "student_full_name", "relation"]:
                parent_info["children"][0][property_name] = value
            else:
                parent_info[property_name] = value
            col_num += 1
        student_code = parent_info["children"][0].get("student_code", "")
        student_full_name = parent_info["children"][0].get("student_full_name", "")
        # ignore blank line
        if not parent_info["full_name"]:
            continue
        # check not null field
        if not parent_info["sex"] or not parent_info["mobile"] or not student_code:
            err_msg = u"第%d行: 用户[姓名/性别/手机号/学生学籍号]为空" % row_num
            ret_data.append(err_msg)
            continue
        # check student is exist
        student_obj = Student.objects.filter(account__code=student_code, del_flag=NO)
        if student_full_name:
            student_obj = student_obj.filter(account__full_name=student_full_name, del_flag=NO)
        if not student_obj:
            err_msg = u"第%d行: 学生不存在" % row_num
            ret_data.append(err_msg)
            continue
        # add parent
        try:
            err_code = add_parent(user, parent_info, is_override)
        except Exception as ex:
            err_code = {"c": -1, "m": ex.message}
        if err_code['c'] == ERR_USER_ALREADY_EXIST[0]:
            skip_id_list.append(row_num)
        elif err_code['c'] != ERR_SUCCESS[0]:
            err_msg = u"第%d行[%s]: " % (row_num, parent_info["full_name"])
            err_msg += err_code['m']
            ret_data.append(err_msg)
        else:
            success_num += 1
    if ret_data:
        ret_msg = u"%d条记录新增成功，%d条记录重复跳过，%d条记录新增失败" % (success_num, len(skip_id_list), len(ret_data))
        return {"c": ERR_IMPORT_DATA_ERROR[0], "m": ret_msg, "d": ret_data}
    else:
        ret_msg = u"%d条记录新增成功，%d条记录重复跳过" % (success_num, len(skip_id_list))
        return {"c": ERR_SUCCESS[0], "m": ret_msg, "d": []}


def export_parent(user, full_name="", student_name="", student_id="", verbose="", is_active=""):
    ret_resp = list_parent(user=user, full_name=full_name, student_name=student_name, student_id=student_id,
                           verbose=verbose, is_active=is_active, display_mode=PARENT_LIST_ONE_PARENT_MULTI_LINE)
    if ret_resp["c"] != ERR_SUCCESS[0]:
        return ret_resp
    parent_list = ret_resp["d"]

    # write to tmp file
    property_display_list = [u"姓名", u"性别", u"手机号", u"身份证号", u"学生姓名", u"学生学籍号", u"称呼"]
    property_name_list = ["full_name", "sex", "mobile", "id_card", "student_full_name", "student_code", "relation"]
    file_path = gen_path()
    wb = openpyxl.Workbook(encoding='utf-8')
    ws = wb.active
    col = 1
    row = 1
    font = Font(bold=True)
    for property_display in property_display_list:
        cell = ws.cell(column=col, row=row, value=property_display)
        cell.font = font
        col += 1
    row += 1
    for parent in parent_list:
        col = 1
        for property_name in property_name_list:
            cell = ws.cell(column=col, row=row, value=parent[property_name])
            col += 1
        row += 1
    reset_excel_column_with(ws)
    wb.save(filename=file_path)

    return {"c": ERR_SUCCESS[0], "m": ERR_SUCCESS[1], "d": [file_path]}


def list_parent_student(user):
    if user.type == USER_TYPE_STUDENT:
        student_obj = Student.objects.filter(school_id=user.school_id, account_id=user.id, del_flag=FLAG_NO).first()
        parent_student_list = ParentStudent.objects.filter(student=student_obj, del_flag=FLAG_NO)
    elif user.type == USER_TYPE_PARENT:
        parent_obj = Parent.objects.filter(school_id=user.school_id, account_id=user.id, del_flag=FLAG_NO).first()
        parent_student_list = ParentStudent.objects.filter(parent=parent_obj, del_flag=FLAG_NO).exclude(status=APPLICATION_STATUS_REFUSED[0])
    else:
        return {"c": ERR_USER_TYPE_ERR[0], "m": ERR_USER_TYPE_ERR[1], "d": []}
    parent_student_list = parent_student_list.values("id", "student_id", "student__full_name", "student__account__code",
                                                     "student__cls__class_name", "parent_id", "parent__full_name",
                                                     "parent__account__mobile", "relation", "comments", "status",
                                                     "student__image__url", "parent__image__url")
    parent_student_list = list(parent_student_list)
    for parent_student in parent_student_list:
        parent_student["student_code"] = parent_student.pop("student__account__code")
        parent_student["student_full_name"] = parent_student.pop("student__full_name")
        parent_student["student_class_name"] = parent_student.pop("student__cls__class_name")
        parent_student["parent_full_name"] = parent_student.pop("parent__full_name")
        parent_student["parent_mobile"] = parent_student.pop("parent__account__mobile")
        parent_student["student_image_url"] = get_image_url(parent_student.pop("student__image__url"))
        parent_student["parent_image_url"] = get_image_url(parent_student.pop("parent__image__url"))

    return {"c": ERR_SUCCESS[0], "m": ERR_SUCCESS[1], "d": parent_student_list}


@transaction.atomic
def delete_parent_student(user, parent_student_id):
    if user.type == USER_TYPE_STUDENT:
        student_obj = Student.objects.filter(school_id=user.school_id, account_id=user.id, del_flag=FLAG_NO).first()
        parent_student_list = ParentStudent.objects.filter(student=student_obj, del_flag=FLAG_NO)
    elif user.type == USER_TYPE_PARENT:
        parent_obj = Parent.objects.filter(school_id=user.school_id, account_id=user.id, del_flag=FLAG_NO).first()
        parent_student_list = ParentStudent.objects.filter(parent=parent_obj, del_flag=FLAG_NO)
    else:
        return {"c": ERR_USER_TYPE_ERR[0], "m": ERR_USER_TYPE_ERR[1], "d": []}
    parent_student_obj = parent_student_list.filter(id=int(parent_student_id)).first()
    if not parent_student_obj:
        return {"c": ERR_SUCCESS[0], "m": ERR_SUCCESS[1], "d": []}
    else:
        parent_student_obj.del_flag = FLAG_YES
        parent_student_obj.save()
    # 检查该家长是否还有学生在该学校
    if not ParentStudent.objects.filter(parent=parent_student_obj.parent, del_flag=FLAG_NO).exists():
        parent_obj = Parent.objects.filter(id=parent_student_obj.parent_id, del_flag=FLAG_NO).first()
        if parent_obj:
            parent_obj.del_flag = FLAG_YES
            parent_obj.save()
    return {"c": ERR_SUCCESS[0], "m": ERR_SUCCESS[1], "d": []}


@transaction.atomic
def update_parent_student(user, parent_student_id, relation, status=""):
    if user.type == USER_TYPE_STUDENT:
        student_obj = Student.objects.filter(school_id=user.school_id, account_id=user.id, del_flag=FLAG_NO).first()
        parent_student_list = ParentStudent.objects.filter(student=student_obj, del_flag=FLAG_NO)
    elif user.type == USER_TYPE_PARENT:
        parent_obj = Parent.objects.filter(school_id=user.school_id, account_id=user.id, del_flag=FLAG_NO).first()
        parent_student_list = ParentStudent.objects.filter(parent=parent_obj, del_flag=FLAG_NO)
    else:
        return {"c": ERR_USER_TYPE_ERR[0], "m": ERR_USER_TYPE_ERR[1], "d": []}
    parent_student_obj = parent_student_list.filter(id=int(parent_student_id)).first()
    if not parent_student_obj:
        return {"c": ERR_SUCCESS[0], "m": ERR_SUCCESS[1], "d": []}
    else:
        if relation:
            parent_student_obj.relation = relation
        if user.type == USER_TYPE_PARENT and status and int(status) in [APPLICATION_STATUS_APPROVED[0], APPLICATION_STATUS_REFUSED[0]]:
            parent_student_obj.status = int(status)
        parent_student_obj.save()
    return {"c": ERR_SUCCESS[0], "m": ERR_SUCCESS[1], "d": []}


@transaction.atomic
def add_parent_by_student(user, mobile, relation, messagecode):
    if messagecode and not check_verify_code(mobile, messagecode):
        return {"c": ERR_VERIFY_CODE_ERROR[0], "m": ERR_VERIFY_CODE_ERROR[1], "d": []}
    if user.type == USER_TYPE_STUDENT:
        student_obj = Student.objects.filter(school_id=user.school_id, account_id=user.id, del_flag=FLAG_NO).first()
        if not student_obj:
            return {"c": ERR_USER_NOT_EXIST[0], "m": ERR_USER_NOT_EXIST[1], "d": []}
    else:
        return {"c": ERR_USER_TYPE_ERR[0], "m": ERR_USER_TYPE_ERR[1], "d": []}

    # 检查该学生是否已经邀请过该家长
    if ParentStudent.objects.filter(student=student_obj, parent__account__mobile=mobile, del_flag=FLAG_NO).exists():
        return {"c": ERR_PARENT_HAVE_EXIST_ERROR[0], "m": ERR_PARENT_HAVE_EXIST_ERROR[1], "d": []}

    # messagecode为空，代表用户已经存在需要发送邀请
    if not messagecode:
        status = APPLICATION_STATUS_NOT_PROCESS[0]
    else:
        status = APPLICATION_STATUS_APPROVED[0]
    parent_full_name = student_obj.full_name + relation
    parent_info = {"mobile": mobile, "full_name": parent_full_name, "children":
        [{"student_id": student_obj.id, "relation": relation, "status": status}]}
    ret_resp = add_parent(user, parent_info)

    return ret_resp

