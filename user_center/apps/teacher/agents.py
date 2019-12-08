# coding=utf-8

from .models import Teacher
from user_center.apps.school.models import Class
from user_center.utils.constant import *
from user_center.apps.account.models import Account
from user_center.utils.err_code import *
from user_center.apps.account.agents import *
from user_center.apps.subject.models import *
from user_center.utils.public_fun import *
from user_center.utils.file_fun import *
from django.db.models import Q
from django.db import transaction
from django.db import models
from user_center.apps.service.models import UserRole


import logging
import openpyxl
from openpyxl.styles import Font

logger = logging.getLogger(__name__)


def list_teacher(user=None, full_name='', code='', name_or_code='', is_in='', in_date_year='',
                 kind='', is_available='', verbose='', school_id='', title=''):
    if school_id:
        school_id = int(school_id)
    else:
        school_id = user.school_id
    teachers_query = Teacher.objects.filter(school_id=school_id, del_flag=FLAG_NO)
    if verbose:
        verbose = int(verbose)
        if verbose == 0:  # 在校
            teachers_query = teachers_query.filter(is_in=FLAG_YES)
        elif verbose == 1:  # 离校
            teachers_query = teachers_query.filter(is_in=FLAG_NO)
        else:  # 所有
            pass
    else:
        teachers_query = teachers_query.filter(is_in=FLAG_YES)
    if full_name:
        teachers_query = teachers_query.filter(full_name__contains=full_name)
    if code:
        teachers_query = teachers_query.filter(Q(school_code__contains=code) | Q(tmp_code__contains=code))
    if name_or_code:
        teachers_query = teachers_query.filter(Q(school_code__contains=name_or_code) |
                                           Q(tmp_code__contains=name_or_code) |
                                           Q(full_name__contains=name_or_code))
    if is_in:
        is_in = str_to_int(is_in)
        if is_in > 0:
            teachers_query = teachers_query.filter(is_in=is_in)
    if kind:
        teachers_query = teachers_query.filter(kind=kind)
    if is_available:
        is_available = str_to_int(is_available)
        if is_available > 0:
            teachers_query = teachers_query.filter(is_available=is_available)
    if in_date_year:
        in_date_year = str_to_int(in_date_year)
        if in_date_year > 0:
            teachers_query = teachers_query.filter(in_date__year=in_date_year)
    if title:
        if title != TILE_NAME_CLASSMASTER:
            teachers_query = teachers_query.filter(title__name=title)
        else:
            teachers_query = teachers_query.filter(cls__isnull=False)

    teachers_data = teachers_query.values('id', 'account_id', 'account__username', 'id_card',
                                          'school_code', 'tmp_code', 'full_name',
                                          'sex', 'email', 'account__mobile', 'image__url',
                                          'account__type','cls_id', 'kind', 'is_in', 'is_available',
                                          'in_date', 'out_date', 'comments', 'title__name', 'account__last_login')

    _data = []
    for teacher in teachers_data:
        if teacher['cls_id']:
            if teacher['title__name']:
                if teacher['title__name'] == TILE_NAME_TEACHER:
                    teacher['title__name'] = ""
                else:
                    teacher['title__name'] += u"、"
            else:
                teacher['title__name'] = ""
            teacher['title__name'] += TILE_NAME_CLASSMASTER
        _data_item = dict(id=str(teacher['id']), account_id=teacher['account_id'], username=teacher['account__username'],
                          id_card=teacher['id_card'], school_code=teacher['school_code'], tmp_code=teacher['tmp_code'],
                          full_name=teacher['full_name'], sex=teacher['sex'], email=teacher['email'],
                          mobile=teacher['account__mobile'], type=teacher['account__type'], cls_id=teacher['cls_id'],
                          kind=teacher['kind'], is_in=teacher['is_in'], is_available=teacher['is_available'],
                          in_date=datetime_to_str(teacher['in_date']), out_date=datetime_to_str(teacher['out_date']), comments=teacher['comments'],
                          title=teacher['title__name'], last_login=datetime_to_str(teacher["account__last_login"], DATE_FORMAT_TIME))
        _data_item['image_url'] = get_image_url(teacher['image__url']) if teacher['image__url'] else ""
        _data.append(_data_item)
    dict_resp = dict(c=ERR_SUCCESS[0], m=ERR_SUCCESS[1], d=_data)
    return dict_resp



"""
keys =['id', 'username', 'id_card', 'code', 'tmp_code',
               'fullname', 'sex', 'email', 'mobile', 'type',
               'cls_id', 'kind', 'is_in', 'is_available', 'in_date', 'out_date', 'comments']
"""


@transaction.atomic
def update_teacher(user, teacher_info, account_id):
    school_id = user.school_id

    teacher_id = teacher_info.get("id", "")
    if not teacher_id:
        return {"c": ERR_USER_NOT_EXIST[0], "m": ERR_USER_NOT_EXIST[1], "d": []}
    teacher = Teacher.objects.filter(id=teacher_id, school_id=school_id, del_flag=NO).first()
    if not teacher:
        return {"c": ERR_USER_NOT_EXIST[0], "m": ERR_USER_NOT_EXIST[1], "d": []}
    # 检查自己修改自己的信息
    if account_id and int(account_id) != teacher.account_id:
        return {"c": ERR_USER_AUTH[0], "m": ERR_USER_AUTH[1], "d": []}

    username = teacher_info.get("username")
    full_name = clean_string(teacher_info.get("full_name"))
    id_card = teacher_info.get("id_card")
    school_code = teacher_info.get("school_code")
    tmp_code = teacher_info.get("tmp_code")
    mobile = teacher_info.get("mobile")
    sex = teacher_info.get("sex")
    email = teacher_info.get("email")
    address = teacher_info.get("address")
    image_url = teacher_info.get("image_url")
    class_id = teacher_info.get("class_id")
    kind = teacher_info.get("kind")
    is_in = teacher_info.get("is_in")
    is_available = teacher_info.get("is_available")
    in_date = teacher_info.get("in_date")
    out_date = teacher_info.get("out_date", "")
    comments = teacher_info.get("comments")
    title = teacher_info.get("title", "")
    birthday = teacher_info.get("birthday")
    native_place = teacher_info.get("native_place")

    image_obj = None
    if image_url:
        image_url = get_file_name_from_url(image_url)
        image_obj = Image.objects.filter(url__contains=image_url).first()
        if not image_obj:
            return {"c": ERR_USER_INFO_IMAGE_NOT_EXIST[0], "m": ERR_USER_INFO_IMAGE_NOT_EXIST[1], "d": []}

    # 检查身份证号是否已经占用
    if id_card and check_id_code_have_occupied(id_card, teacher.account_id):
        raise Exception(ERR_USER_ID_CARD_CONFLICT_ERROR[1])

    # 检查职务名称是否存在
    title_obj = None
    if title:
        title_obj = Title.objects.filter(school_id=school_id, name=title, del_flag=NO).first()
        if not title_obj:
            return {"c": ERR_TITLE_NOT_EXIST_ERROR[0], "m": ERR_TITLE_NOT_EXIST_ERROR[1], "d": []}
    src_title = teacher.title
    src_cls_id = teacher.cls_id
    # update account
    account = teacher.account
    err_code = update_account(user=user, account=account, username=username, mobile=mobile)
    if err_code != ERR_SUCCESS:
        return {"c": err_code[0], "m": err_code[1], "d": []}

    # update teacher
    if full_name: teacher.full_name = full_name
    if id_card: teacher.id_card = id_card
    elif id_card == "": teacher.id_card = id_card
    if sex: teacher.sex = sex
    if email: teacher.email = email
    elif email == "": teacher.email = email
    if address: teacher.address = address
    elif address == "": teacher.address = address
    if image_obj: teacher.image = image_obj
    if kind: teacher.kind = kind
    elif kind == "": teacher.kind = kind
    if class_id: teacher.cls_id = int(class_id)
    if is_in: teacher.is_in = int(is_in)
    if is_available: teacher.is_available = int(is_available)
    if in_date: teacher.in_date = str_to_datetime(in_date)
    if out_date: teacher.out_date = str_to_datetime(out_date)
    if comments: teacher.comments = comments
    elif comments == "": teacher.comments = comments
    if title_obj: teacher.title = title_obj
    if birthday: teacher.birthday = str_to_datetime(birthday)
    if native_place: teacher.native_place = native_place
    if school_code:
        if school_code == teacher.school_code:
            pass
        else:
            if Teacher.objects.filter(school_id=school_id, del_flag=FLAG_NO, school_code=school_code).exists():
                raise Exception(u"教工号已存在,操作失败")
            teacher.school_code = school_code
    elif school_code == "": 
        teacher.school_code = school_code
    if tmp_code: teacher.tmp_code = tmp_code
    elif tmp_code == "": teacher.tmp_code = tmp_code

    teacher.save()
    check_and_delete_account(account.id)
    if title_obj and src_title != title:
        src_title_obj = Title.objects.filter(school_id=school_id, name=title, del_flag=NO).first()
        refresh_teacher_amount(school_id, [src_title_obj, title_obj])
    elif title_obj and src_cls_id != class_id:
        src_title_obj = Title.objects.filter(school_id=school_id, name=TILE_NAME_CLASSMASTER, del_flag=NO).first()
        refresh_teacher_amount(school_id, [src_title_obj])
    return {"c": ERR_SUCCESS[0], "m": ERR_SUCCESS[1], "d": []}


@transaction.atomic
def add_teacher(user, teacher_data, is_override=False):
    school_id = user.school_id

    username = teacher_data.get("username", "")
    password = teacher_data.get("password", "")
    full_name = clean_string(teacher_data.get("full_name", ""))
    id_card = teacher_data.get("id_card", "")
    school_code = teacher_data.get("school_code", "")
    tmp_code = teacher_data.get("tmp_code", "")
    sex = teacher_data.get("sex", "")
    mobile = teacher_data.get("mobile", "")
    email = teacher_data.get("email", "")
    # type = teacher_data.get("type", 0)
    # if type: type = int(type)
    address = teacher_data.get("address", "")
    # company = teacher_data.get("company", "")
    title = teacher_data.get("title", "")
    is_mobile_login = teacher_data.get("is_mobile_login", "") or 0
    if is_mobile_login: is_mobile_login = int(is_mobile_login)
    is_email_login = teacher_data.get("is_email_login", "") or 0
    if is_email_login: is_email_login = int(is_email_login)
    image_url = teacher_data.get("image_url", "")
    image_obj = None
    if image_url:
        image_url = get_file_name_from_url(image_url)
        image_obj = Image.objects.filter(url__contains=image_url).first()
        if not image_obj:
            return {"c": ERR_USER_INFO_IMAGE_NOT_EXIST[0], "m": ERR_USER_INFO_IMAGE_NOT_EXIST[1], "d": []}

    # check information integrity
    if (not mobile and not username) or not full_name:
        logger.error(" teacher info is not integrity")
        return {"c": ERR_USER_INFO_INCOMPLETE[0], "m": ERR_USER_INFO_INCOMPLETE[1], "d": []}

    if school_code:
        if Teacher.objects.filter(school_id=school_id, del_flag=FLAG_NO, school_code=school_code).exists():
            raise Exception(u"教工号已存在,操作失败")
    # 检查帐号在同一个学校的同一个角色是否已经存在
    if Teacher.objects.filter(school_id=school_id, account__mobile=mobile, del_flag=NO).exists():
        return {"c": ERR_USER_ALREADY_EXIST[0], "m": ERR_USER_ALREADY_EXIST[1], "d": []}
    # 检查职务名称是否存在
    if title:
        title_obj = Title.objects.filter(school_id=school_id, name=title, del_flag=NO).first()
        if not title_obj:
            return {"c": ERR_TITLE_NOT_EXIST_ERROR[0], "m": ERR_TITLE_NOT_EXIST_ERROR[1], "d": []}
    else:
        title_obj = Title.objects.filter(school_id=school_id, name=TILE_NAME_TEACHER, del_flag=NO).first()
    # add an new account
    if username:
        err_code = add_account(username=username, password=password, type=USER_TYPE_TEACHER, school_id=school_id)
    else:
        err_code = add_account(mobile=mobile, password=password, type=USER_TYPE_TEACHER, school_id=school_id)
    if err_code[0] != ERR_SUCCESS[0]:
        return {"c": err_code[0], "m": err_code[1], "d": []}
    account_id = err_code[1]

    # 检查身份证号是否已经占用
    if id_card and check_id_code_have_occupied(id_card, int(account_id)):
        raise Exception(ERR_USER_ID_CARD_CONFLICT_ERROR[1])

    # find class
    class_name = teacher_data.get("class_name", "")
    cls_id = teacher_data.get("cls_id", "")
    cls = None
    if cls_id:
        cls = Class.objects.filter(school_id=school_id, id=cls_id, del_flag=NO).first()
        if not cls:
            return {"c": ERR_CLASS_NOT_EXIST[0], "m": ERR_CLASS_NOT_EXIST[1], "d": []}
    elif class_name:
        cls = Class.objects.filter(school_id=school_id, class_name=class_name, graduate_status=NO, del_flag=NO).first()
        if not cls:
            return {"c": ERR_CLASS_NOT_EXIST[0], "m": ERR_CLASS_NOT_EXIST[1], "d": []}

    # add teacher
    kind = teacher_data.get("kind", "")
    is_in = teacher_data.get("is_in", "")
    is_available = teacher_data.get("is_available", "")
    in_date = teacher_data.get("in_date", "")
    out_date = teacher_data.get("out_date", "")
    comments = teacher_data.get("comments", "")
    birthday = teacher_data.get("birthday", "")
    native_place = teacher_data.get("native_place", "")

    teacher = Teacher.objects.create(school_id=school_id, account_id=account_id)

    if full_name: teacher.full_name=full_name
    if id_card: teacher.id_card = id_card
    if school_code: teacher.school_code = school_code
    if tmp_code: teacher.tmp_code = tmp_code
    if sex: teacher.sex = sex
    if email: teacher.email = email
    if image_obj: teacher.image = image_obj
    if kind: teacher.kind = kind
    if cls: teacher.cls = cls
    if is_in:
        teacher.is_in = int(range_yes_no(is_in))
    if is_available:
        teacher.is_available = int(range_yes_no(is_available))
    if in_date: teacher.in_date = str_to_datetime(in_date)
    if out_date: teacher.out_date = str_to_datetime(out_date)
    if address: teacher.address = address
    if comments: teacher.comments = comments
    if title_obj: teacher.title = title_obj
    if birthday: teacher.birthday = str_to_datetime(birthday)
    if native_place: teacher.native_place = native_place
    teacher.save()
    if title_obj:
        refresh_teacher_amount(school_id, [title_obj])
    return {"c": ERR_SUCCESS[0], "m": ERR_SUCCESS[1], "d": [teacher.id]}


def import_teacher(user, file_obj, is_override=False):
    school_id = user.school_id
    ret_data = []
    skip_id_list = []
    success_num = 0

    if not file_obj.name.endswith('xlsx'):
        return {"c": ERR_FILE_FORMAT_NOT_SUPPORTED[0], "m": ERR_FILE_FORMAT_NOT_SUPPORTED[1],  "d": []}

    # write file data to tmp path
    file_path = gen_path()
    with open(file_path, 'wb+') as destination:
        for chunk in file_obj.chunks():
            destination.write(chunk)

    wb = openpyxl.load_workbook(file_path)
    ws = wb.get_active_sheet()

    property_display_list = [u"姓名*", u"性别*", u"手机号*", u"教工号", u"管理班级名称（班主任）", u"身份证号"]
    property_name_list = ["full_name", "sex", "mobile", "school_code", "class_name", "id_card"]
    property_len = len(property_name_list)

    # check the first row
    for i in range(property_len):
        if property_display_list[i] != ws.cell(column=i+1, row=1).value:
            return {"c": ERR_FILE_TEMPLATE_ERROR[0], "m": ERR_FILE_TEMPLATE_ERROR[1],  "d": []}
    row_num = 0

    for row in ws.rows:
        row_num += 1
        if row_num == 1:
            # 第一行为标题栏
            continue
        col_num = 0
        teacher_info = {}
        for cell in row:
            if col_num >= property_len:
                break
            property_name = property_name_list[col_num]
            value = ""
            if isinstance(cell.value, (int, long)):
                value = str(cell.value)
            elif isinstance(cell.value, basestring):
                value = cell.value.strip()
            teacher_info[property_name] = value
            col_num += 1
        # ignore blank line
        if not teacher_info["full_name"]:
            continue
        # check not null field
        if not teacher_info["sex"] or not teacher_info["mobile"]:
            err_msg = u"第%d行: 用户[姓名/性别/手机号]为空" % row_num
            ret_data.append(err_msg)
            continue
        try:
            err_code = add_teacher(user, teacher_info, is_override)
        except Exception as ex:
            err_code = {"c": -1, "m": ex.message}
        if err_code['c'] == ERR_USER_ALREADY_EXIST[0]:
            skip_id_list.append(row_num)
        elif err_code['c'] != ERR_SUCCESS[0]:
            err_msg = u"第%d行[%s]: " % (row_num, teacher_info["full_name"])
            err_msg += err_code['m']
            ret_data.append(err_msg)
        else:
            success_num += 1
    if ret_data:
        ret_msg = u"%d条记录新增成功，%d条记录重复跳过，%d条记录新增失败" % (success_num, len(skip_id_list), len(ret_data))
        return {"c": ERR_IMPORT_DATA_ERROR[0], "m": ret_msg, "d": ret_data}
    else:
        ret_msg = u"%d条记录新增成功，%d条记录重复跳过" % (success_num, len(skip_id_list))
        return {"c": ERR_SUCCESS[0], "m": ret_msg, "d": []}


def export_teacher(user, full_name='', code='', name_or_code='', is_in='', in_date_year='',
                   kind='', is_available='', verbose=''):
    dict_resp = list_teacher(user=user, full_name=full_name, code=code, name_or_code=name_or_code,
                             is_in=is_in, in_date_year=in_date_year, kind=kind, is_available=is_available,
                             verbose=verbose)
    if dict_resp['c'] != ERR_SUCCESS[0]:
        return dict_resp
    teacher_list = dict_resp['d']

    # write to tmp file
    property_display_list = [u"姓名", u"性别", u"手机号", u"教工号", u"邮箱", u"身份证号", u"入校时间"]
    property_name_list = ["full_name", "sex", "mobile", "school_code", "email", "id_card", "in_date"]
    file_path = gen_path()
    wb = openpyxl.Workbook(encoding='utf-8')
    ws = wb.active
    col = 1
    row = 1
    font = Font(bold=True)
    for property_display in property_display_list:
        cell = ws.cell(column=col, row=row, value=property_display)
        cell.font = font
        col += 1
    row += 1
    for teacher in teacher_list:
        col = 1
        for property_name in property_name_list:
            cell = ws.cell(column=col, row=row, value=teacher[property_name])
            col += 1
        row += 1
    reset_excel_column_with(ws)
    wb.save(filename=file_path)

    return {"c": ERR_SUCCESS[0], "m": ERR_SUCCESS[1], "d": [file_path]}


def detail_teacher(id, account_id):
    teacher_data = Teacher.objects.filter(id=int(id), del_flag=FLAG_NO).values('id', 'account_id', 'account__username',
                                                                          'full_name', 'id_card', 'school__name_full', 'school__name_simple',
                                                                          'school_code', 'tmp_code', 'sex',
                                                                          'account__mobile', 'email', 'address',
                                                                          'kind', 'is_in', 'is_available', 'in_date', 'out_date',
                                                                          'comments', 'image__url', 'cls_id',
                                                                          'title__name', 'birthday', 'native_place').first()

    # 检查自己修改自己的信息
    if account_id and int(account_id) != teacher_data["account_id"]:
        return {"c": ERR_USER_AUTH[0], "m": ERR_USER_AUTH[1], "d": []}

    # teacher_data = dict(teacher_data)

    # if teacher_data['cls_id']:
    #     if teacher_data['title__name']:
    #         teacher_data['title__name'] += u"、"
    #     else:
    #         teacher_data['title__name'] = ""
    #     teacher_data['title__name'] += TILE_NAME_CLASSMASTER
    _data = dict(id=teacher_data['id'], account_id=teacher_data['account_id'], username=teacher_data['account__username'],
                 full_name=teacher_data['full_name'], id_card=teacher_data['id_card'],
                 school_code=teacher_data['school_code'], tmp_code=teacher_data['tmp_code'], address=teacher_data['address'],
                 mobile=teacher_data['account__mobile'], email=teacher_data['email'], sex=teacher_data['sex'],
                 kind=teacher_data['kind'], is_in=teacher_data['is_in'], is_available=teacher_data['is_available'],
                 in_date=datetime_to_str(teacher_data['in_date']), out_date=datetime_to_str(teacher_data['out_date']), comments=teacher_data['comments'],
                 title=teacher_data['title__name'], birthday=datetime_to_str(teacher_data['birthday']), native_place=teacher_data['native_place'],
                 school_name_full=teacher_data['school__name_full'],
                 school_name_simple=teacher_data['school__name_simple'] if teacher_data['school__name_simple'] else '',
                 )
    _data['image_url'] = get_image_url(teacher_data['image__url']) if teacher_data['image__url'] else ""
    return dict(c=ERR_SUCCESS[0], m=ERR_SUCCESS[1], d=[_data])


@transaction.atomic
def delete_teacher(user, id_list):
    school_id = user.school_id
    teacher_id_list = map(lambda x: int(x), id_list)
    delet_filter = Teacher.objects.filter(school_id=school_id, id__in=teacher_id_list, del_flag=FLAG_NO).\
        exclude(account_id=user.id).values_list('id', flat=True)
    delet_filter = list(delet_filter)
    undeletable = UserRole.objects.filter(user_id__in=delet_filter, del_flag=FLAG_NO).values_list('user_id', flat=True)
    undeletable = list(undeletable)
    deletable = list(set(delet_filter) - set(undeletable))
    del_count = len(deletable)
    Teacher.objects.filter(id__in=deletable, del_flag=FLAG_NO).update(del_flag=FLAG_YES, update_time=datetime.datetime.now())
    account_id_list = list(Teacher.objects.filter(id__in=deletable).values_list('account_id', flat=True))

    # 删除对应的account对象
    delete_list=[]
    for id in account_id_list:
        if not account_ref_status(id):
            delete_list.append(id)

    Account.objects.filter(id__in=delete_list, del_flag=FLAG_NO).update(del_flag=FLAG_YES, update_time=datetime.datetime.now())
    if del_count > 0:
        refresh_teacher_amount(school_id)
    if len(teacher_id_list) == del_count:
        return dict(c=ERR_SUCCESS[0], m=ERR_SUCCESS[1], d=[])
    else:
        msg = u"[%d个删除成功%d个删除失败:不能删除自身账号或者请先删除老师的管理员角色]" % (del_count, len(teacher_id_list)-del_count)
        return dict(c=ERR_DELETE_PART[0], m=ERR_DELETE_PART[1] + msg, d=[])


@transaction.atomic
def leave_teacher(user, teacher_id_list, is_leave=""):
    school_id = user.school_id
    teacher_id_list = map(lambda x: int(x), teacher_id_list)

    total_num = len(teacher_id_list)
    real_num = 0
    for teacher_id in teacher_id_list:
        teacher_obj = Teacher.objects.filter(school_id=school_id, id=teacher_id, del_flag=FLAG_NO).first()
        if not teacher_obj:
            continue
        if is_leave and int(is_leave) == 0:
            teacher_obj.is_in = True
        else:
            teacher_obj.is_in = False
            teacher_obj.cls = None
        teacher_obj.save()
        real_num += 1
    refresh_teacher_amount(school_id=school_id)
    # return msg
    if real_num == total_num:
        return {"c": ERR_SUCCESS[0], "m": ERR_SUCCESS[1], "d": []}
    elif total_num == 1:
        return {"c": ERR_OP_ERROR[0], "m": ERR_OP_ERROR[1], "d": []}
    else:
        msg = u"%d条记录操作失败，%d条记录操作成功" % (total_num-real_num, real_num)
        return {"c": ERR_OP_PART_ERROR[0], "m": msg, "d": []}


def account_ref_status(id):
    account = Account.objects.get(id=id)
    return account.teacher_set.filter(del_flag=FLAG_NO).exists() or account.parent_account.filter(del_flag=FLAG_NO).\
        exists() or account.student_set.filter(del_flag=FLAG_NO).exists()


def list_teacher_class(user, teacher_id=""):
    if teacher_id and ADMIN_USER_TYPE_SYSTEM_ADMIN not in user.role and ADMIN_USER_TYPE_TEACHER_ADMIN not in user.role:
        return {"c": ERR_USER_AUTH[0], "m": ERR_USER_AUTH[1], "d": []}
    if teacher_id:
        teacher_id = int(teacher_id)
    else:
        teacher_id = Teacher.objects.filter(account_id=user.id, school_id=user.school_id, del_flag=FLAG_NO).first()
    class_list = TeacherClass.objects.filter(teacher_id=teacher_id, cls__graduate_status=FLAG_NO, del_flag=FLAG_NO).\
        values("cls_id", "cls__class_name", "cls__class_alias", "is_master")
    ret_class_list = []
    for class_info in class_list:
        ret_class_info = {"class_id": class_info["cls_id"],
                          "class_name": class_info["cls__class_name"],
                          "class_alias": class_info["cls__class_alias"],
                          "is_master": class_info["is_master"],}
        ret_class_list.append(ret_class_info)
    return {"c": ERR_SUCCESS[0], "m": ERR_SUCCESS[1], "d": ret_class_list}


def add_teacher_class(user, teacher_id, class_id_list):
    if teacher_id and ADMIN_USER_TYPE_SYSTEM_ADMIN not in user.role and ADMIN_USER_TYPE_TEACHER_ADMIN not in user.role:
        return {"c": ERR_USER_AUTH[0], "m": ERR_USER_AUTH[1], "d": []}
    if teacher_id:
        teacher_id = int(teacher_id)
    else:
        teacher_obj = Teacher.objects.filter(account_id=user.id, school_id=user.school_id, del_flag=FLAG_NO).first()
        if teacher_obj:
            teacher_id = teacher_obj.id
        else:
            return {"c": ERR_USER_AUTH[0], "m": ERR_USER_AUTH[1], "d": []}
    class_id_list = json.loads(class_id_list)
    for class_id in class_id_list:
        obj, created = TeacherClass.objects.update_or_create(teacher_id=teacher_id, cls_id=class_id, del_flag=FLAG_NO)
        if not created:
            obj.update_time = datetime.datetime.now()
            obj.save()
    # TODO sync to other service
    return {"c": ERR_SUCCESS[0], "m": ERR_SUCCESS[1], "d": []}


@transaction.atomic
def update_teacher_master_class(teacher_id_list, class_id):
    teacher_id_list = map(lambda x: int(x), teacher_id_list)
    TeacherClass.objects.filter(cls_id=class_id, is_master=FLAG_YES, del_flag=FLAG_NO).exclude(teacher_id__in=teacher_id_list).\
        update(del_flag=FLAG_YES, update_time=datetime.datetime.now())
    for teacher_id in teacher_id_list:
        TeacherClass.objects.filter(teacher_id=teacher_id, is_master=FLAG_YES, del_flag=FLAG_NO).exclude(cls_id=class_id)\
            .update(del_flag=FLAG_YES, update_time=datetime.datetime.now())

        obj, created = TeacherClass.objects.update_or_create(teacher_id=teacher_id, cls_id=class_id, del_flag=FLAG_NO)
        obj.is_master = FLAG_YES
        obj.update_time = datetime.datetime.now()
        obj.save()


def init_teacher_master_class(school_id):
    teacher_list = Teacher.objects.filter(school_id=school_id, cls__graduate_status=FLAG_NO, del_flag=FLAG_NO)
    for teacher in teacher_list:
        obj, created = TeacherClass.objects.update_or_create(teacher_id=teacher.id, cls_id=teacher.cls_id, del_flag=FLAG_NO)
        obj.is_master = FLAG_YES
        obj.update_time = datetime.datetime.now()
        obj.save()


def delete_teacher_class(user, teacher_id, class_id_list):

    if teacher_id and ADMIN_USER_TYPE_SYSTEM_ADMIN not in user.role and ADMIN_USER_TYPE_TEACHER_ADMIN not in user.role:
        return {"c": ERR_USER_AUTH[0], "m": ERR_USER_AUTH[1], "d": []}
    if teacher_id:
        teacher_id = int(teacher_id)
    else:
        teacher_id = Teacher.objects.filter(account_id=user.id, school_id=user.school_id, del_flag=FLAG_NO).first()
    class_id_list = json.loads(class_id_list)

    total_num = len(class_id_list)
    real_num = 0
    for class_id in class_id_list:
        rows = TeacherClass.objects.filter(cls_id=class_id, teacher_id=teacher_id, is_master=FLAG_NO, del_flag=FLAG_NO).\
            update(del_flag=FLAG_YES, update_time=datetime.datetime.now())
        real_num += rows

    # return msg
    if real_num == total_num:
        return {"c": ERR_SUCCESS[0], "m": ERR_SUCCESS[1], "d": []}
    elif total_num == 1:
        if TeacherClass.objects.filter(cls_id=class_id_list[0], teacher_id=teacher_id, is_master=FLAG_YES, del_flag=FLAG_NO):
            return {"c": ERR_TEACHER_CLASS_MASTER_DELETE_ERROR[0], "m": ERR_TEACHER_CLASS_MASTER_DELETE_ERROR[1], "d": []}
        return {"c": ERR_DELETE_ERROR[0], "m": ERR_DELETE_ERROR[1], "d": []}
    else:
        msg = u"%d条记录删除失败，%d条记录删除成功" % (total_num-real_num, real_num)
        return {"c": ERR_DELETE_PART[0], "m": msg, "d": []}


def list_teacher_textbook(user, teacher_id="", subject_id=""):
    if teacher_id and ADMIN_USER_TYPE_SYSTEM_ADMIN not in user.role and ADMIN_USER_TYPE_TEACHER_ADMIN not in user.role:
        return {"c": ERR_USER_AUTH[0], "m": ERR_USER_AUTH[1], "d": []}
    if teacher_id:
        teacher_id = int(teacher_id)
    else:
        teacher_id = Teacher.objects.filter(account_id=user.id, school_id=user.school_id, del_flag=FLAG_NO).first()

    school_textbook_id_list = SchoolTextbook.objects.filter(school_id=user.school_id, del_flag=FLAG_NO,
                                                            textbook__is_active=FLAG_YES, textbook__del_flag=FLAG_NO,
                                                            textbook__subject__is_active=FLAG_YES).\
        values_list("textbook_id", flat=True)
    textbook_list = TeacherTextbook.objects.filter(teacher_id=teacher_id,
                                                   textbook_id__in=school_textbook_id_list,
                                                   del_flag=FLAG_NO)

    if subject_id:
        textbook_list = textbook_list.filter(textbook__subject_id=int(subject_id))
    textbook_list = textbook_list.values("textbook_id", "textbook__name", "textbook__subject__name", "is_current")
    textbook_list = list(textbook_list)
    ret_textbook_list = []
    for textbook_info in textbook_list:
        ret_textbook_info = {"textbook_id": textbook_info["textbook_id"],
                             "textbook_name": textbook_info["textbook__name"],
                             "is_current": textbook_info["is_current"],
                             "subject_name": textbook_info["textbook__subject__name"]}
        ret_textbook_list.append(ret_textbook_info)
    return {"c": ERR_SUCCESS[0], "m": ERR_SUCCESS[1], "d": ret_textbook_list}


def add_teacher_textbook(user, teacher_id, textbook_id_list):
    if teacher_id and ADMIN_USER_TYPE_SYSTEM_ADMIN not in user.role and ADMIN_USER_TYPE_TEACHER_ADMIN not in user.role:
        return {"c": ERR_USER_AUTH[0], "m": ERR_USER_AUTH[1], "d": []}
    if teacher_id:
        teacher_id = int(teacher_id)
    else:
        teacher_obj = Teacher.objects.filter(account_id=user.id, school_id=user.school_id, del_flag=FLAG_NO).first()
        teacher_id = teacher_obj.id
    textbook_id_list = json.loads(textbook_id_list)
    for textbook_id in textbook_id_list:
        obj, created = TeacherTextbook.objects.update_or_create(teacher_id=teacher_id, textbook_id=textbook_id, del_flag=FLAG_NO)
        if not created:
            obj.update_time = datetime.datetime.now()
            obj.save()
    __refresh_teacher_textbook(teacher_id)
    # TODO sync to other service
    return {"c": ERR_SUCCESS[0], "m": ERR_SUCCESS[1], "d": []}


def delete_teacher_textbook(user, teacher_id, textbook_id_list):
    if teacher_id and ADMIN_USER_TYPE_SYSTEM_ADMIN not in user.role and ADMIN_USER_TYPE_TEACHER_ADMIN not in user.role:
        return {"c": ERR_USER_AUTH[0], "m": ERR_USER_AUTH[1], "d": []}
    if teacher_id:
        teacher_id = int(teacher_id)
    else:
        teacher_id = Teacher.objects.filter(account_id=user.id, school_id=user.school_id, del_flag=FLAG_NO).first()
    textbook_id_list = json.loads(textbook_id_list)
    for textbook_id in textbook_id_list:
        TeacherTextbook.objects.filter(textbook_id=textbook_id, teacher_id=teacher_id).\
            update(del_flag=FLAG_YES, update_time=datetime.datetime.now())
    __refresh_teacher_textbook(teacher_id)
    # TODO sync to other service
    return {"c": ERR_SUCCESS[0], "m": ERR_SUCCESS[1], "d": []}


# 更新教师的主授教材
@transaction.atomic()
def update_teacher_textbook(user, teacher_id, textbook_id):
    if teacher_id and ADMIN_USER_TYPE_SYSTEM_ADMIN not in user.role and ADMIN_USER_TYPE_TEACHER_ADMIN not in user.role:
        return {"c": ERR_USER_AUTH[0], "m": ERR_USER_AUTH[1], "d": []}
    if teacher_id:
        teacher_id = int(teacher_id)
    else:
        teacher_id = Teacher.objects.filter(account_id=user.id, school_id=user.school_id, del_flag=FLAG_NO).first()
    # 取消原来的主教材
    TeacherTextbook.objects.filter(teacher_id=teacher_id, is_current=FLAG_YES, del_flag=FLAG_NO).\
        update(is_current=FLAG_NO, update_time=datetime.datetime.now())
    # 新增主教材
    teacher_textbook = TeacherTextbook.objects.filter(textbook_id=int(textbook_id), teacher_id=teacher_id, del_flag=FLAG_NO).first()
    if not teacher_textbook:
        raise Exception(ERR_TEXTBOOK_ID_ERROR[1])
    teacher_textbook.is_current = FLAG_YES
    teacher_textbook.save()

    return {"c": ERR_SUCCESS[0], "m": ERR_SUCCESS[1], "d": []}


# 重新选择一个主授教材
def __refresh_teacher_textbook(teacher_id, textbook_id=None):
    if TeacherTextbook.objects.filter(teacher_id=teacher_id, is_current=FLAG_YES, del_flag=FLAG_NO).exists():
        return
    if not textbook_id:
        teacher_textbook_obj = TeacherTextbook.objects.filter(teacher_id=teacher_id, del_flag=FLAG_NO).first()
        if not teacher_textbook_obj:
            return
        textbook_id = teacher_textbook_obj.textbook_id
    teacher_textbook = TeacherTextbook.objects.filter(textbook_id=textbook_id, teacher_id=teacher_id, del_flag=FLAG_NO).first()
    teacher_textbook.is_current = FLAG_YES
    teacher_textbook.save()


# 更新职务中的教师总数
def refresh_teacher_amount(school_id, title_obj_list=None):
    if not title_obj_list:
        title_obj_list = Title.objects.filter(school_id=school_id, del_flag=FLAG_NO)
    for title_obj in title_obj_list:
        if title_obj.name == TILE_NAME_CLASSMASTER:
            teacher_amount = Teacher.objects.filter(school_id=school_id, is_in=True, cls__isnull=False, del_flag=FLAG_NO).count()
        else:
            teacher_amount = Teacher.objects.filter(school_id=school_id, is_in=True, title=title_obj, del_flag=FLAG_NO).count()
        title_obj.teacher_amount = teacher_amount
        title_obj.save()


def init_teacher_title(school_id):
    title_obj = Title.objects.filter(school_id=school_id, name=TILE_NAME_TEACHER, del_flag=FLAG_NO).first()
    Teacher.objects.filter(school_id=school_id, title__isnull=True).update(title=title_obj)

