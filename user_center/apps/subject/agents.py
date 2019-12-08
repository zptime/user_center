# -*- coding=utf-8 -*-

from django.db.models import F
from django.db import transaction
import re

import openpyxl
from openpyxl.styles import Font

from models import *
from user_center.utils.file_fun import *
from user_center.utils.public_fun import *


logger = logging.getLogger(__name__)

GRADE_NUM_NAME_MAPPING = {GRADE_NUM_FIRST[0]:GRADE_NUM_FIRST[1], GRADE_NUM_SECOND[0]:GRADE_NUM_SECOND[1], GRADE_NUM_THIRD[0]:GRADE_NUM_THIRD[1],
                          GRADE_NUM_FOURTH[0]:GRADE_NUM_FOURTH[1], GRADE_NUM_FIFTH[0]:GRADE_NUM_FIFTH[1], GRADE_NUM_SIXTH[0]:GRADE_NUM_SIXTH[1],
                          GRADE_NUM_SEVENTH[0]:GRADE_NUM_SEVENTH[1], GRADE_NUM_EIGHTH[0]:GRADE_NUM_EIGHTH[1], GRADE_NUM_NINTH[0]:GRADE_NUM_NINTH[1],
                          GRADE_NUM_TENTH[0]:GRADE_NUM_TENTH[1], GRADE_NUM_ELEVENTH[0]:GRADE_NUM_ELEVENTH[1], GRADE_NUM_TWELFTH[0]:GRADE_NUM_TWELFTH[1]
                          }

@transaction.atomic
def admin_add_subject(user, subject_name=""):
    account_id = user.id
    subject_name = subject_name.strip()

    if not subject_name:
        return {"c": -1, "m": u"科目名称不能为空", "d": []}

    if Subject.objects.filter(name=subject_name, del_flag=0).exists():
        return {"c": ERR_SUBJECT_HAVE_EXIST_ERROR[0], "m": ERR_SUBJECT_HAVE_EXIST_ERROR[1], "d": []}

    subject = Subject.objects.create(name=subject_name, editor_id=account_id)

    subject.save()

    return {"c": ERR_SUCCESS[0], "m": ERR_SUCCESS[1], "d": []}


def admin_list_subject(user, is_active="1", subject_name=""):
    is_active = int(is_active)

    subject_list = Subject.objects.filter(is_active=is_active, del_flag=0).order_by("-update_time").all()

    if subject_name:
        subject_list = subject_list.filter(name__contains=subject_name)

    subject_info_list = subject_list.values("id", "name", "editor__full_name", "update_time")

    ret_subject_list = []
    for subject_info in subject_info_list:
        ret_subject = dict(  id=str(subject_info["id"]),
                             name=subject_info["name"],
                             editor_name=subject_info["editor__full_name"],
                             update_time=datetime_to_str(subject_info["update_time"],DATE_FORMAT_TIME)
                          )
        ret_subject_list.append(ret_subject)

    dict_resp = {"c": ERR_SUCCESS[0], "m": ERR_SUCCESS[1], "d": ret_subject_list}
    return dict_resp


@transaction.atomic
def admin_edit_subject(user, subject_id="", subject_name=""):
    account_id = user.id
    subject_name = subject_name.strip()

    if not subject_id or not subject_name:
        return {"c": ERR_REQUEST_PARAMETER_ERROR[0], "m": ERR_REQUEST_PARAMETER_ERROR[1], "d": []}

    subject_id = int(subject_id)

    subject_obj = Subject.objects.filter(id=subject_id, is_active=1, del_flag=0).first()

    if not subject_obj:
        return {"c": ERR_SUBJECT_ID_ERROR[0], "m": ERR_SUBJECT_ID_ERROR[1], "d": []}

    if subject_obj.name != subject_name:
        if Subject.objects.filter(name=subject_name, del_flag=0).exists():
            return {"c": ERR_SUBJECT_HAVE_EXIST_ERROR[0], "m": ERR_SUBJECT_HAVE_EXIST_ERROR[1], "d": []}

        subject_obj.name = subject_name
        subject_obj.editor_id = account_id

    subject_obj.save()
    dict_resp = {"c": ERR_SUCCESS[0], "m": ERR_SUCCESS[1], "d": []}
    return dict_resp


@transaction.atomic
def admin_freeze_subject(user, subject_id_list=[]):
    if not subject_id_list:
        return {"c": ERR_SUBJECT_ID_ERROR[0], "m": ERR_SUBJECT_ID_ERROR[1], "d": []}

    subject_id_list = map(lambda x : int(x), subject_id_list)

    for subject_id in subject_id_list:
        if not Subject.objects.filter(id=subject_id, is_active=1, del_flag=0).exists():
            return {"c": ERR_SUBJECT_ID_ERROR[0], "m": ERR_SUBJECT_ID_ERROR[1], "d": []}

    Subject.objects.filter(id__in=subject_id_list, del_flag=0).update(is_active=0, update_time=datetime.datetime.now())

    # 将关联的教材也冻结
    Textbook.objects.filter(subject_id__in=subject_id_list, del_flag=0).update(is_active=0, update_time=datetime.datetime.now())

    dict_resp = {"c": ERR_SUCCESS[0], "m": ERR_SUCCESS[1], "d": []}
    return dict_resp


@transaction.atomic
def admin_unfreeze_subject(user, subject_id_list=[]):
    if not subject_id_list:
        return {"c": ERR_SUBJECT_ID_ERROR[0], "m": ERR_SUBJECT_ID_ERROR[1], "d": []}

    subject_id_list = map(lambda x : int(x), subject_id_list)

    for subject_id in subject_id_list:
        if not Subject.objects.filter(id=subject_id, is_active=0, del_flag=0).exists():
            return {"c": ERR_SUBJECT_ID_ERROR[0], "m": ERR_SUBJECT_ID_ERROR[1], "d": []}

    Subject.objects.filter(id__in=subject_id_list, del_flag=0).update(is_active=1, update_time=datetime.datetime.now())

    # 将关联的教材也解冻
    Textbook.objects.filter(subject_id__in=subject_id_list, del_flag=0).update(is_active=1, update_time=datetime.datetime.now())

    dict_resp = {"c": ERR_SUCCESS[0], "m": ERR_SUCCESS[1], "d": []}
    return dict_resp


@transaction.atomic
def admin_delete_subject(user, subject_id_list=[]):
    if not subject_id_list:
        return {"c": ERR_SUBJECT_ID_ERROR[0], "m": ERR_SUBJECT_ID_ERROR[1], "d": []}

    subject_id_list = map(lambda x : int(x), subject_id_list)

    for subject_id in subject_id_list:
        if not Subject.objects.filter(id=subject_id, is_active=0, del_flag=0).exists():
            return {"c": ERR_SUBJECT_ID_ERROR[0], "m": ERR_SUBJECT_ID_ERROR[1], "d": []}

    Subject.objects.filter(id__in=subject_id_list, del_flag=0).update(del_flag=1, update_time=datetime.datetime.now())
    # 将学校的科目同步删除
    SchoolSubject.objects.filter(subject_id__in=subject_id_list, del_flag=0).update(del_flag=1, update_time=datetime.datetime.now())
    # 将关联的教材也删除
    Textbook.objects.filter(subject_id__in=subject_id_list, del_flag=0).update(del_flag=1, update_time=datetime.datetime.now())
    # 将学校的教材同步删除
    SchoolTextbook.objects.filter(textbook__subject_id__in=subject_id_list, del_flag=0).update(del_flag=1, update_time=datetime.datetime.now())

    dict_resp = {"c": ERR_SUCCESS[0], "m": ERR_SUCCESS[1], "d": []}
    return dict_resp


def admin_import_subject(user, file_obj=""):
    account_id = user.id

    if not file_obj.name.endswith('xlsx'):
        return {"c": ERR_FILE_FORMAT_NOT_SUPPORTED[0], "m": ERR_FILE_FORMAT_NOT_SUPPORTED[1],  "d": []}

    ret_data = []
    skip_id_list = []
    success_num = 0

    # write file data to tmp path
    file_path = gen_path()
    with open(file_path, 'wb+') as destination:
        for chunk in file_obj.chunks():
            destination.write(chunk)

    wb = openpyxl.load_workbook(filename=file_path, read_only=True)
    ws = wb.get_active_sheet()

    property_display_list = [u"科目"]
    property_name_list = ["subject_name"]
    property_len = len(property_name_list)

    # check the first row
    for i in range(property_len):
        try:
            if not re.search(property_display_list[i], ws.cell(column=i+1, row=1).value):
                return {"c": ERR_FILE_TEMPLATE_ERROR[0], "m": ERR_FILE_TEMPLATE_ERROR[1],  "d": []}
        except:
            return {"c": ERR_FILE_TEMPLATE_ERROR[0], "m": ERR_FILE_TEMPLATE_ERROR[1],  "d": []}

    # 检查表格数据是否有误，若有误，则返回错误数据提示信息
    # row_num = 0
    # for row in ws.rows:
    #     row_num += 1
    #     if row_num == 1:
    #         # 第一行为标题栏
    #         continue
    #     col_num = 0
    #     subject_info = {}
    #     for cell in row:
    #         if col_num >= property_len:
    #             break
    #         property_name = property_name_list[col_num]
    #         value = ""
    #         if isinstance(cell.value, (int, long)):
    #             value = str(cell.value)
    #         elif isinstance(cell.value, basestring):
    #             value = cell.value.strip()
    #         subject_info[property_name] = value
    #         col_num += 1
    #     # 忽略空白行
    #     if not subject_info["subject_name"]:
    #         continue
    #
    #     if Subject.objects.filter(name=subject_info["subject_name"], del_flag=0).exists():
    #         err_msg = u"第%d行: 科目名称已经存在" % row_num
    #         ret_data.append(err_msg)
    #
    # if ret_data:
    #     return {"c": -1, "m": u"表格数据有误，数据没有导入，请修改后重新导入", "d": ret_data}

    # 如果表格数据无误，则导入课程数据
    row_num = 0
    for row in ws.rows:
        row_num += 1
        if row_num == 1:
            # 第一行为标题栏
            continue
        col_num = 0
        subject_info = {}
        for cell in row:
            if col_num >= property_len:
                break
            property_name = property_name_list[col_num]
            value = ""
            if isinstance(cell.value, (int, long)):
                value = str(cell.value)
            elif isinstance(cell.value, basestring):
                value = cell.value.strip()
            subject_info[property_name] = value
            col_num += 1
        # 忽略空白行
        if not subject_info["subject_name"]:
            continue

        err_code = admin_add_subject(user=user, subject_name=subject_info["subject_name"])
        if err_code['c'] == ERR_SUBJECT_HAVE_EXIST_ERROR[0]:
            skip_id_list.append(row_num)
        elif err_code['c'] != ERR_SUCCESS[0]:
            err_msg = u"第%d行 [%s]: " % (row_num, subject_info["subject_name"])
            err_msg += err_code['m']
            ret_data.append(err_msg)
        else:
            success_num += 1

    if ret_data:
        ret_msg = u"%d条记录新增成功，%d条记录重复跳过，%d条记录新增失败" % (success_num, len(skip_id_list), len(ret_data))
        return {"c": ERR_IMPORT_DATA_ERROR[0], "m": ret_msg, "d": ret_data}
    else:
        ret_msg = u"%d条记录新增成功，%d条记录重复跳过" % (success_num, len(skip_id_list))
        return {"c": ERR_SUCCESS[0], "m": ret_msg, "d": []}


def admin_export_subject(user, is_active="1"):
    is_active = int(is_active)

    subject_name_list = list(Subject.objects.filter(is_active=is_active, del_flag=0).values_list("name", flat=True))

    # write to tmp file
    property_display_list = [u"科目"]
    # property_name_list = ["subject_name"]
    file_path = gen_path()
    wb = openpyxl.Workbook(encoding='utf-8')
    ws = wb.active
    col = 1
    row = 1
    font = Font(bold=True)
    for property_display in property_display_list:
        cell = ws.cell(column=col, row=row, value=property_display)
        cell.font = font
        col += 1
    row += 1
    for subject_name in subject_name_list:
        col = 1
        ws.cell(column=col, row=row, value=subject_name)
        row += 1
    reset_excel_column_with(ws)
    wb.save(filename=file_path)

    return {"c": ERR_SUCCESS[0], "m": ERR_SUCCESS[1], "d": [file_path]}


@transaction.atomic
def school_add_subject(user, subject_id_list=[]):
    school_id = user.school_id

    if not subject_id_list:
        return {"c": ERR_SUBJECT_ID_ERROR[0], "m": ERR_SUBJECT_ID_ERROR[1], "d": []}

    subject_id_list = map(lambda x : int(x), subject_id_list)

    for subject_id in subject_id_list:
        if not Subject.objects.filter(id=subject_id, is_active=1, del_flag=0).exists():
            return {"c": ERR_SUBJECT_ID_ERROR[0], "m": ERR_SUBJECT_ID_ERROR[1], "d": []}

    if SchoolSubject.objects.filter(school_id=school_id, subject_id__in=subject_id_list, del_flag=0).exists():
        return {"c": ERR_SUBJECT_ID_ERROR[0], "m": ERR_SUBJECT_ID_ERROR[1], "d": []}
    else:
        for subject_id in subject_id_list:
            SchoolSubject.objects.create(school_id=school_id, subject_id=subject_id)

    return {"c": ERR_SUCCESS[0], "m": ERR_SUCCESS[1], "d": []}


def school_check_subject(user, subject_id=""):
    school_id = user.school_id
    if not subject_id:
        return {"c": ERR_SUBJECT_ID_ERROR[0], "m": ERR_SUBJECT_ID_ERROR[1], "d": []}

    #if not SchoolSubject.objects.filter(school_id=school_id, subject_id=subject_id, del_flag=0).exists():
    #    return {"c": ERR_SUBJECT_ID_ERROR[0], "m": ERR_SUBJECT_ID_ERROR[1], "d": []}

    if SchoolTextbook.objects.filter(school_id=school_id, textbook__subject_id=subject_id, textbook__is_active=1, del_flag=0).exists():
        return {"c": ERR_SUCCESS[0], "m": ERR_SUCCESS[1], "d": ["True"]}
    else:
        return {"c": ERR_SUCCESS[0], "m": ERR_SUCCESS[1], "d": ["False"]}


def school_list_subject(user):
    school_id = user.school_id

    subject_info_list = SchoolSubject.objects.filter(school_id=school_id, subject__is_active=1, subject__del_flag=0, del_flag=0)\
                                                           .order_by("-update_time").values("subject_id", "subject__name")

    ret_subject_list = []
    for subject_info in subject_info_list:
        ret_subject = dict(  id=str(subject_info["subject_id"]),
                             name=subject_info["subject__name"],
                          )
        ret_subject_list.append(ret_subject)

    dict_resp = {"c": ERR_SUCCESS[0], "m": ERR_SUCCESS[1], "d": ret_subject_list}
    return dict_resp


@transaction.atomic
def school_update_subject(user, subject_id_list=[]):
    school_id = user.school_id
    if subject_id_list:
        subject_id_list = map(lambda x: int(x), subject_id_list)

    exist_subject_id_list = list(SchoolSubject.objects.filter(school_id=school_id, subject__is_active=1, del_flag=0)
                                    .values_list("subject_id", flat=True))

    delete_subject_id_list = list(set(exist_subject_id_list).difference(set(subject_id_list)))
    add_subject_id_list = list(set(subject_id_list).difference(set(exist_subject_id_list)))

    # 检查要删除的科目id是否正确
    for subject_id in delete_subject_id_list:
        if not SchoolSubject.objects.filter(school_id=school_id, subject_id=subject_id, del_flag=0).exists():
            return {"c": ERR_SUBJECT_ID_ERROR[0], "m": ERR_SUBJECT_ID_ERROR[1], "d": []}

    # 检查要删除科目下是否存在教材
    err_code = []
    for subject_id in delete_subject_id_list:
        if SchoolTextbook.objects.filter(school_id=school_id, textbook__subject_id=subject_id, textbook__is_active=1, del_flag=0).exists():
            subject_obj = Subject.objects.filter(id=subject_id, is_active=1, del_flag=0).first()
            subject_name = subject_obj.name
            err_messege = u"科目 [%s] 下存在相关教材，无法删除" % (subject_name)
            err_code.append(err_messege)
    if err_code:
        err_code.append(u"操作失败")
        return {"c": ERR_SUCCESS[0], "m": ERR_SUCCESS[1], "d": err_code}

    # 删除相关科目和教材
    SchoolSubject.objects.filter(school_id=school_id, subject_id__in=delete_subject_id_list, del_flag=0).update(del_flag=1, update_time=datetime.datetime.now())
    SchoolTextbook.objects.filter(school_id=school_id, textbook__subject_id__in=delete_subject_id_list, del_flag=0).update(del_flag=1, update_time=datetime.datetime.now())

    for subject_id in add_subject_id_list:
        if not Subject.objects.filter(id=subject_id, is_active=1, del_flag=0).exists():
            return {"c": ERR_SUBJECT_ID_ERROR[0], "m": ERR_SUBJECT_ID_ERROR[1], "d": []}

    if SchoolSubject.objects.filter(school_id=school_id, subject_id__in=add_subject_id_list, del_flag=0).exists():
        return {"c": ERR_SUBJECT_ID_ERROR[0], "m": ERR_SUBJECT_ID_ERROR[1], "d": []}
    else:
        for subject_id in add_subject_id_list:
            SchoolSubject.objects.create(school_id=school_id, subject_id=subject_id)

    return {"c": ERR_SUCCESS[0], "m": ERR_SUCCESS[1], "d": []}


@transaction.atomic
def school_delete_subject(user, subject_id_list=[]):
    school_id = user.school_id

    if not subject_id_list:
        return {"c": ERR_SUBJECT_ID_ERROR[0], "m": ERR_SUBJECT_ID_ERROR[1], "d": []}

    subject_id_list = map(lambda x : int(x), subject_id_list)

    for subject_id in subject_id_list:
        if not SchoolSubject.objects.filter(school_id=school_id, subject_id=subject_id, del_flag=0).exists():
            return {"c": ERR_SUBJECT_ID_ERROR[0], "m": ERR_SUBJECT_ID_ERROR[1], "d": []}

    # 检查要删除科目下是否存在教材
    err_code = []
    for subject_id in subject_id_list:
        if SchoolTextbook.objects.filter(school_id=school_id, textbook__subject_id=subject_id, textbook__is_active=1, del_flag=0).exists():
            subject_obj = Subject.objects.filter(id=subject_id, is_active=1, del_flag=0).first()
            subject_name = subject_obj.name
            err_messege = u"科目 [%s] 下存在相关教材，无法删除" % (subject_name)
            err_code.append(err_messege)
    if err_code:
        err_code.append(u"操作失败")
        return {"c": ERR_SUCCESS[0], "m": ERR_SUCCESS[1], "d": err_code}

    SchoolSubject.objects.filter(school_id=school_id, subject_id__in=subject_id_list, del_flag=0).update(del_flag=1, update_time=datetime.datetime.now())
    SchoolTextbook.objects.filter(school_id=school_id, textbook__subject_id__in=subject_id_list, del_flag=0).update(del_flag=1, update_time=datetime.datetime.now())

    dict_resp = {"c": ERR_SUCCESS[0], "m": ERR_SUCCESS[1], "d": []}
    return dict_resp


@transaction.atomic
def admin_add_textbook(user, subject_id="", textbook_name="", grade_num=""):
    account_id = user.id
    textbook_name = textbook_name.strip()

    if not subject_id or not textbook_name or not grade_num:
        return {"c": ERR_REQUEST_PARAMETER_ERROR[0], "m": ERR_REQUEST_PARAMETER_ERROR[1], "d": []}

    subject_id = int(subject_id)
    grade_num = int(grade_num)

    if not Subject.objects.filter(id=subject_id, is_active=1, del_flag=0).exists():
        return {"c": ERR_SUBJECT_ID_ERROR[0], "m": ERR_SUBJECT_ID_ERROR[1], "d": []}

    if not (0 < grade_num < 13):
        return {"c": ERR_SUBJECT_GRADE_NUM_ERROR[0], "m": ERR_SUBJECT_GRADE_NUM_ERROR[1], "d": []}

    if Textbook.objects.filter(subject_id=subject_id, name=textbook_name, grade_num=grade_num, del_flag=0).exists():
        return {"c": ERR_TEXTBOOK_HAVE_EXIST_ERROR[0], "m": ERR_TEXTBOOK_HAVE_EXIST_ERROR[1], "d": []}

    Textbook.objects.create(subject_id=subject_id, editor_id=account_id, name=textbook_name, grade_num=grade_num)

    dict_resp = {"c": ERR_SUCCESS[0], "m": ERR_SUCCESS[1], "d": []}
    return dict_resp


def admin_list_textbook(user, is_active="1", subject_id_list=[], grade_num_list=[]):
    account_id = user.id
    is_active = int(is_active)
    if subject_id_list:
        subject_id_list = map(lambda x: int(x), subject_id_list)
    if grade_num_list:
        grade_num_list = map(lambda x: int(x), grade_num_list)

    textbook_list = Textbook.objects.filter(is_active=is_active, del_flag=0).all()

    if subject_id_list:
        textbook_list = textbook_list.filter(subject_id__in=subject_id_list)

    if grade_num_list:
        textbook_list = textbook_list.filter(grade_num__in=grade_num_list)

    textbook_info_list = textbook_list.order_by("grade_num", "subject_id")\
                                        .values("id", "grade_num", "subject__name", "name", "chapter_count")

    ret_textbook_list = []
    for textbook_info in textbook_info_list:
        ret_textbook = dict( id=str(textbook_info["id"]),
                             grade_num=str(textbook_info["grade_num"]),
                             grade_name=GRADE_NUM_NAME_MAPPING[textbook_info["grade_num"]],
                             subject_name=textbook_info["subject__name"],
                             textbook_name=textbook_info["name"],
                             chapter_count=textbook_info["chapter_count"]
                           )
        ret_textbook_list.append(ret_textbook)

    dict_resp = {"c": ERR_SUCCESS[0], "m": ERR_SUCCESS[1], "d": ret_textbook_list}
    return dict_resp


def admin_detail_textbook(user, textbook_id=""):
    if not textbook_id:
        return {"c": ERR_REQUEST_PARAMETER_ERROR[0], "m": ERR_REQUEST_PARAMETER_ERROR[1], "d": []}
    textbook_id=int(textbook_id)

    textbook_obj = Textbook.objects.filter(id=textbook_id, is_active=1, del_flag=0)\
                                              .values("id", "grade_num", "subject__name", "name", "chapter_count").first()
    if not textbook_obj:
        return {"c": ERR_TEXTBOOK_ID_ERROR[0], "m": ERR_TEXTBOOK_ID_ERROR[1], "d": []}

    ret_textbook = dict(  id=str(textbook_obj["id"]),
                          grade_num=str(textbook_obj["grade_num"]),
                          grade_name=GRADE_NUM_NAME_MAPPING[textbook_obj["grade_num"]],
                          subject_name=textbook_obj["subject__name"],
                          textbook_name=textbook_obj["name"],
                          chapter_count=textbook_obj["chapter_count"]
                        )

    dict_resp = {"c": ERR_SUCCESS[0], "m": ERR_SUCCESS[1], "d": [ret_textbook]}
    return dict_resp


@transaction.atomic
def admin_edit_textbook(user, textbook_id="", textbook_name="", grade_num=""):
    account_id = user.id
    textbook_name = textbook_name.strip()

    if not textbook_id or not textbook_name or not grade_num:
        return {"c": ERR_REQUEST_PARAMETER_ERROR[0], "m": ERR_REQUEST_PARAMETER_ERROR[1], "d": []}

    grade_num = int(grade_num)

    textbook_obj = Textbook.objects.filter(id=textbook_id, is_active=1, del_flag=0).first()

    if not textbook_obj:
        return {"c": ERR_TEXTBOOK_ID_ERROR[0], "m": ERR_TEXTBOOK_ID_ERROR[1], "d": []}

    if textbook_obj.name != textbook_name or textbook_obj.grade_num != grade_num:
        if Textbook.objects.filter(name=textbook_name, grade_num=grade_num, del_flag=0).exists():
            return {"c": ERR_TEXTBOOK_HAVE_EXIST_ERROR[0], "m": ERR_TEXTBOOK_HAVE_EXIST_ERROR[1], "d": []}

        textbook_obj.name = textbook_name
        textbook_obj.grade_num = grade_num
        textbook_obj.editor_id = account_id

        textbook_obj.save()

    dict_resp = {"c": ERR_SUCCESS[0], "m": ERR_SUCCESS[1], "d": []}
    return dict_resp


@transaction.atomic
def admin_freeze_textbook(user, textbook_id=""):
    account_id = user.id

    if not textbook_id:
        return {"c": ERR_TEXTBOOK_ID_ERROR[0], "m": ERR_TEXTBOOK_ID_ERROR[1], "d": []}

    textbook_obj = Textbook.objects.filter(id=textbook_id, is_active=1, del_flag=0).first()

    if not textbook_obj:
        return {"c": ERR_TEXTBOOK_ID_ERROR[0], "m": ERR_TEXTBOOK_ID_ERROR[1], "d": []}

    textbook_obj.is_active = 0

    textbook_obj.save()

    dict_resp = {"c": ERR_SUCCESS[0], "m": ERR_SUCCESS[1], "d": []}
    return dict_resp


@transaction.atomic
def admin_unfreeze_textbook(user, textbook_id=""):
    account_id = user.id

    if not textbook_id:
        return {"c": ERR_TEXTBOOK_ID_ERROR[0], "m": ERR_TEXTBOOK_ID_ERROR[1], "d": []}

    textbook_obj = Textbook.objects.filter(id=textbook_id, is_active=0, del_flag=0).first()

    if not textbook_obj:
        return {"c": ERR_TEXTBOOK_ID_ERROR[0], "m": ERR_TEXTBOOK_ID_ERROR[1], "d": []}

    textbook_obj.is_active = 1

    textbook_obj.save()

    dict_resp = {"c": ERR_SUCCESS[0], "m": ERR_SUCCESS[1], "d": []}
    return dict_resp


@transaction.atomic
def admin_delete_textbook(user, textbook_id=""):
    school_id = user.school_id

    if not textbook_id:
        return {"c": ERR_TEXTBOOK_ID_ERROR[0], "m": ERR_TEXTBOOK_ID_ERROR[1], "d": []}

    textbook_obj = Textbook.objects.filter(id=textbook_id, is_active=0, del_flag=0).first()

    if not textbook_obj:
        return {"c": ERR_TEXTBOOK_ID_ERROR[0], "m": ERR_TEXTBOOK_ID_ERROR[1], "d": []}

    textbook_obj.del_flag = 1
    # 同步删除学校的教材
    SchoolTextbook.objects.filter(textbook_id=textbook_id, del_flag=0).update(del_flag=1, update_time=datetime.datetime.now())
    # 同步删除教材的章节
    Chapter.objects.filter(textbook_id=textbook_id, del_flag=0).update(del_flag=1, update_time=datetime.datetime.now())

    textbook_obj.save()

    dict_resp = {"c": ERR_SUCCESS[0], "m": ERR_SUCCESS[1], "d": []}
    return dict_resp


@transaction.atomic
def school_add_textbook(user, textbook_id=""):
    school_id = user.school_id

    if not textbook_id:
        return {"c": ERR_REQUEST_PARAMETER_ERROR[0], "m": ERR_REQUEST_PARAMETER_ERROR[1], "d": []}

    textbook_id = int(textbook_id)

    if not Textbook.objects.filter(id=textbook_id, is_active=1, del_flag=0).exists():
        return {"c": ERR_TEXTBOOK_ID_ERROR[0], "m": ERR_TEXTBOOK_ID_ERROR[1], "d": []}

    if SchoolTextbook.objects.filter(school_id=school_id, textbook_id=textbook_id, del_flag=0).exists():
        return {"c": ERR_TEXTBOOK_HAVE_EXIST_ERROR[0], "m": ERR_TEXTBOOK_HAVE_EXIST_ERROR[1], "d": []}
    else:
        SchoolTextbook.objects.create(school_id=school_id, textbook_id=textbook_id)

    dict_resp = {"c": ERR_SUCCESS[0], "m": ERR_SUCCESS[1], "d": []}
    return dict_resp


def school_list_textbook(user, subject_id_list=[], grade_num_list=[], not_belong_flag="0"):
    school_id = user.school_id
    not_belong_flag = int(not_belong_flag)
    if subject_id_list:
        subject_id_list = map(lambda x: int(x), subject_id_list)
    if grade_num_list:
        grade_num_list = map(lambda x: int(x), grade_num_list)

    textbook_id_list = SchoolTextbook.objects.filter(school_id=school_id, del_flag=0).values_list("textbook_id", flat=True)
    textbook_id_list = list(textbook_id_list)

    if not_belong_flag:
        school_subject_id_list = SchoolSubject.objects.filter(school_id=school_id, subject__is_active=1, subject__del_flag=0, del_flag=0).values_list("subject_id", flat=True)
        admin_textbook_id_list = Textbook.objects.filter(subject_id__in=school_subject_id_list, is_active=1, del_flag=0).values_list("id", flat=True)
        admin_textbook_id_list = list(admin_textbook_id_list)
        textbook_id_list = list(set(admin_textbook_id_list).difference(set(textbook_id_list)))

    textbook_list = Textbook.objects.filter(id__in=textbook_id_list, is_active=1, subject__is_active=1, del_flag=0).all()

    if grade_num_list:
        textbook_list = textbook_list.filter(grade_num__in=grade_num_list)

    if subject_id_list:
        textbook_list = textbook_list.filter(subject_id__in=subject_id_list)

    textbook_info_list = textbook_list.order_by("grade_num", "subject_id").values("id", "grade_num", "subject_id", "subject__name", "name")

    ret_textbook_list = []
    for textbook_info in textbook_info_list:
        ret_textbook = dict( id=str(textbook_info["id"]),
                             grade_num=str(textbook_info["grade_num"]),
                             grade_name=GRADE_NUM_NAME_MAPPING[textbook_info["grade_num"]],
                             subject_id=textbook_info["subject_id"],
                             subject_name=textbook_info["subject__name"],
                             textbook_name=textbook_info["name"],
                           )
        ret_textbook_list.append(ret_textbook)

    dict_resp = {"c": ERR_SUCCESS[0], "m": ERR_SUCCESS[1], "d": ret_textbook_list}
    return dict_resp


@transaction.atomic
def school_delete_textbook(user, textbook_id=""):
    school_id = user.school_id

    if not textbook_id:
        return {"c": ERR_TEXTBOOK_ID_ERROR[0], "m": ERR_TEXTBOOK_ID_ERROR[1], "d": []}

    if not Textbook.objects.filter(id=textbook_id, is_active=1, del_flag=0).exists():
        return {"c": ERR_TEXTBOOK_ID_ERROR[0], "m": ERR_TEXTBOOK_ID_ERROR[1], "d": []}

    if not SchoolTextbook.objects.filter(school_id=school_id, textbook_id=textbook_id, del_flag=0).exists():
        return {"c": ERR_TEXTBOOK_ID_ERROR[0], "m": ERR_TEXTBOOK_ID_ERROR[1], "d": []}

    SchoolTextbook.objects.filter(school_id=school_id, textbook_id=textbook_id, del_flag=0).update(del_flag=1, update_time=datetime.datetime.now())

    dict_resp = {"c": ERR_SUCCESS[0], "m": ERR_SUCCESS[1], "d": []}
    return dict_resp


@transaction.atomic
def admin_add_chapter(user, textbook_id="", chapter_name="", parent_id=None, sn=""):
    account_id = user.id
    chapter_name = chapter_name.strip()

    if not textbook_id or not chapter_name or not sn:
        return {"c": ERR_REQUEST_PARAMETER_ERROR[0], "m": ERR_REQUEST_PARAMETER_ERROR[1], "d": []}

    sn = int(sn)
    textbook_id = int(textbook_id)
    if not Textbook.objects.filter(id=textbook_id, is_active=1, del_flag=0).exists():
        return {"c": ERR_TEXTBOOK_ID_ERROR[0], "m": ERR_TEXTBOOK_ID_ERROR[1], "d": []}

    if parent_id is None:
        return {"c": ERR_REQUEST_PARAMETER_ERROR[0], "m": ERR_REQUEST_PARAMETER_ERROR[1], "d": []}

    if parent_id != "":
        parent_id = int(parent_id)
        if parent_id and (not Chapter.objects.filter(id=parent_id, textbook_id=textbook_id, del_flag=0).exists()):
            return {"c": ERR_REQUEST_PARAMETER_ERROR[0], "m": ERR_REQUEST_PARAMETER_ERROR[1], "d": []}

    chapter = Chapter.objects.create(textbook_id=textbook_id, name=chapter_name)

    if parent_id:
        Chapter.objects.filter(textbook_id=textbook_id, parent_id=parent_id, sn__gte=sn, del_flag=0).update(sn=F("sn")+1, update_time=datetime.datetime.now())
        chapter.sn = sn

        # 设置父节点id
        chapter.parent_id = parent_id
    else:
        Chapter.objects.filter(textbook_id=textbook_id, parent_id__isnull=True, sn__gte=sn, del_flag=0).update(sn=F("sn")+1, update_time=datetime.datetime.now())
        chapter.sn = sn

    chapter.save()

    #更新教材的章节数目
    chapter_count = Chapter.objects.filter(textbook_id=textbook_id, del_flag=0).count()
    Textbook.objects.filter(id=textbook_id, is_active=1, del_flag=0).update(chapter_count=chapter_count, update_time=datetime.datetime.now())

    dict_resp = {"c": ERR_SUCCESS[0], "m": ERR_SUCCESS[1], "d": []}
    return dict_resp


def admin_list_chapter(user, textbook_id=""):
    if not textbook_id:
        return {"c": ERR_TEXTBOOK_ID_ERROR[0], "m": ERR_TEXTBOOK_ID_ERROR[1], "d": []}

    textbook_id = int(textbook_id)

    textbook_obj = Textbook.objects.filter(id=textbook_id, is_active=1, del_flag=0).first()

    if not textbook_obj:
        return {"c": ERR_TEXTBOOK_ID_ERROR[0], "m": ERR_TEXTBOOK_ID_ERROR[1], "d": []}

    chapter_info_list = list(Chapter.objects.filter(textbook_id=textbook_id, del_flag=0).order_by("parent_id", "-sn")\
                                 .values("id", "parent_id", "sn", "name"))

    #将章节的ID按照关系排序
    chapter_id_list = []
    for chapter_info in chapter_info_list:
        if chapter_info["parent_id"] is None:
            index = 0
        else:
            index = chapter_id_list.index(chapter_info["parent_id"]) + 1
        chapter_id_list.insert(index, chapter_info["id"])


    ret_chapter_list = []
    for chapter_id in chapter_id_list:
        for chapter_info in chapter_info_list:
            if chapter_id == chapter_info["id"]:
                if chapter_info["parent_id"] is None:
                    chapter_info["parent_id"] = 0
                ret_chapter = dict(  id=chapter_info["id"],
                                     pId=chapter_info["parent_id"],
                                     name=chapter_info["name"],
                                     sn=chapter_info["sn"]
                                  )
                ret_chapter_list.append(ret_chapter)
                chapter_info_list.remove(chapter_info)
                break

    dict_resp = {"c": ERR_SUCCESS[0], "m": ERR_SUCCESS[1], "d": ret_chapter_list}
    return dict_resp


@transaction.atomic
def admin_edit_chapter(user, chapter_id="", chapter_name=""):
    chapter_name = chapter_name.strip()

    if not chapter_id or not chapter_name:
        return {"c": ERR_REQUEST_PARAMETER_ERROR[0], "m": ERR_REQUEST_PARAMETER_ERROR[1], "d": []}

    chapter_id = int(chapter_id)

    chapter_obj = Chapter.objects.filter(id=chapter_id, del_flag=0).first()
    if not chapter_obj:
        return {"c": ERR_CHAPTER_ID_ERROR[0], "m": ERR_CHAPTER_ID_ERROR[1], "d": []}

    chapter_obj.name = chapter_name

    chapter_obj.save()

    dict_resp = {"c": ERR_SUCCESS[0], "m": ERR_SUCCESS[1], "d": []}
    return dict_resp


@transaction.atomic
def admin_move_chapter(user, chapter_id="", parent_id=None, sn=""):
    account_id = user.id

    if not chapter_id or not sn:
        return {"c": ERR_REQUEST_PARAMETER_ERROR[0], "m": ERR_REQUEST_PARAMETER_ERROR[1], "d": []}

    sn = int(sn)
    chapter_id = int(chapter_id)

    if parent_id is None:
        return {"c": ERR_REQUEST_PARAMETER_ERROR[0], "m": ERR_REQUEST_PARAMETER_ERROR[1], "d": []}

    if parent_id != "":
        parent_id = int(parent_id)
        if parent_id and (not Chapter.objects.filter(id=parent_id, del_flag=0).exists()):
            return {"c": ERR_REQUEST_PARAMETER_ERROR[0], "m": ERR_REQUEST_PARAMETER_ERROR[1], "d": []}

    chapter_obj = Chapter.objects.filter(id=chapter_id, del_flag=0).first()
    if not chapter_obj:
        return {"c": ERR_TEXTBOOK_ID_ERROR[0], "m": ERR_TEXTBOOK_ID_ERROR[1], "d": []}

    if parent_id:
        Chapter.objects.filter(textbook_id=chapter_obj.textbook_id, parent_id=chapter_obj.parent_id, sn__gt=chapter_obj.sn, del_flag=0)\
                               .update(sn=F("sn")-1, update_time=datetime.datetime.now())
        Chapter.objects.filter(textbook_id=chapter_obj.textbook_id, parent_id=parent_id, sn__gte=sn, del_flag=0)\
                               .update(sn=F("sn")+1, update_time=datetime.datetime.now())

        # 设置父节点id
        chapter_obj.parent_id = parent_id
    else:
        Chapter.objects.filter(textbook_id=chapter_obj.textbook_id, parent_id=chapter_obj.parent_id, sn__gt=chapter_obj.sn, del_flag=0)\
                               .update(sn=F("sn")-1, update_time=datetime.datetime.now())
        Chapter.objects.filter(textbook_id=chapter_obj.textbook_id, parent_id__isnull=True, sn__gte=sn, del_flag=0)\
                               .update(sn=F("sn")+1, update_time=datetime.datetime.now())

    chapter_obj.sn = sn

    chapter_obj.save()

    #更新教材的章节数目
    chapter_count = Chapter.objects.filter(textbook_id=chapter_obj.textbook_id, del_flag=0).count()
    Textbook.objects.filter(id=chapter_obj.textbook_id, is_active=1, del_flag=0).update(chapter_count=chapter_count, update_time=datetime.datetime.now())

    dict_resp = {"c": ERR_SUCCESS[0], "m": ERR_SUCCESS[1], "d": []}
    return dict_resp


@transaction.atomic
def admin_delete_chapter(user, chapter_id=""):
    account_id = user.id

    if not chapter_id:
        return {"c": ERR_REQUEST_PARAMETER_ERROR[0], "m": ERR_REQUEST_PARAMETER_ERROR[1], "d": []}

    chapter_id = int(chapter_id)

    chapter_obj = Chapter.objects.filter(id=chapter_id, del_flag=0).first()
    if not chapter_obj:
        return {"c": ERR_TEXTBOOK_ID_ERROR[0], "m": ERR_TEXTBOOK_ID_ERROR[1], "d": []}

    parent_id = chapter_obj.parent_id

    Chapter.objects.filter(textbook_id=chapter_obj.textbook_id, parent_id=parent_id, sn__gt=chapter_obj.sn, del_flag=0)\
                            .update(sn=F("sn")-1, update_time=datetime.datetime.now())

    def delete_children_node(parent_id):
        children_id_list = list(Chapter.objects.filter(parent_id=parent_id, del_flag=0).values_list("id", flat=True))
        for id in children_id_list:
            Chapter.objects.filter(id=id, del_flag=0).update(del_flag=1, update_time=datetime.datetime.now())
            delete_children_node(id)

    delete_children_node(chapter_obj.id)

    chapter_obj.del_flag = 1
    chapter_obj.save()

    #更新教材的章节数目
    chapter_count = Chapter.objects.filter(textbook_id=chapter_obj.textbook_id, del_flag=0).count()
    Textbook.objects.filter(id=chapter_obj.textbook_id, is_active=1, del_flag=0).update(chapter_count=chapter_count, update_time=datetime.datetime.now())

    dict_resp = {"c": ERR_SUCCESS[0], "m": ERR_SUCCESS[1], "d": []}
    return dict_resp