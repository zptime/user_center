# -*- coding=utf-8 -*-

from django.db.models import Q
from models import *
from user_center.utils.constant import *
from user_center.utils.err_code import *
from user_center.utils.file_fun import get_image_url, gen_path
from user_center.utils.public_fun import *
from user_center.apps.account.models import *
from user_center.apps.teacher.models import *
import openpyxl
from openpyxl.styles import Font
from django.conf import settings
from django.db import transaction
from django.db.utils import ProgrammingError
import json
import traceback


def __convert_role_name(role_name):
    if not role_name:
        return ""
    else:
        role_name_list = role_name.split("-")
        return role_name_list[-1]


def __convert_service_name(service_name):
    if not service_name:
        return ""
    else:
        role_name_list = service_name.split("-")
        return role_name_list[0]


def list_admin_role(user, teacher_id):
    school_id = user.school_id
    # 获取该学校开通的所用服务的角色
    service_id_list = SchoolService.objects.filter(school_id=school_id, del_flag=NO).values_list("service_id", flat=True)
    role_list = Role.objects.filter(service_id__in=list(service_id_list), del_flag=NO).order_by("service_id").values('service__name', 'id', 'name')
    # 获取该教师已经赋予的角色列表
    already_role_id_list = []
    if teacher_id:
        user_role_id_list = UserRole.objects.filter(school_id=user.school_id, user_id=int(teacher_id),
                                                    del_flag=FLAG_NO).values_list("role_id", flat=True)
        already_role_id_list = list(user_role_id_list)
    # 转换输出格式
    ret_role_list = []
    for role in role_list:
        ret_role = {"service_name": __convert_service_name(role["service__name"]), "role_id": role["id"],
                    "role_name": __convert_role_name(role["name"]), "is_already": FLAG_NO}
        if role["id"] in already_role_id_list:
            ret_role["is_already"] = FLAG_YES
        ret_role_list.append(ret_role)
    return {"c": ERR_SUCCESS[0], "m": ERR_SUCCESS[1], "d": ret_role_list}


def list_admin_user(user, service_name="", user_name=""):
    school_id = user.school_id
    # 获取该学校开通的所用服务的角色
    service_id_list = SchoolService.objects.filter(school_id=school_id, del_flag=NO).values_list("service_id", flat=True)
    role_list = Role.objects.filter(service_id__in=list(service_id_list), del_flag=NO).order_by("service_id").all()
    if service_name:
        role_list = role_list.filter(service__name__contains=service_name, del_flag=NO)
    role_id_list = role_list.values_list('id', flat=True)

    # 获取这些角色对应的用户
    admin_user_list = UserRole.objects.filter(school_id=school_id, role_id__in=list(role_id_list), user__is_in=YES, del_flag=NO)
    if user_name:
        admin_user_list = admin_user_list.filter(user__full_name__contains=user_name)
    admin_user_list = admin_user_list.values('id', 'role_id', 'role__name', 'role__service__name', 'user_id',
                                             'user__full_name', 'user__account__mobile', 'user__account__last_login')

    # 转换输出格式
    ret_admin_user_list = []
    for admin_user in admin_user_list:
        ret_admin_user = {"user_role_id": admin_user["id"], "user_id": admin_user["user_id"],
                          "user_name": admin_user["user__full_name"], "mobile": admin_user["user__account__mobile"],
                          "service_name": __convert_service_name(admin_user["role__service__name"]), "role_id": admin_user["role_id"],
                          "role_name": __convert_role_name(admin_user["role__name"]),
                          "last_login": datetime_to_str(admin_user["user__account__last_login"], DATE_FORMAT_TIME)}
        ret_admin_user_list.append(ret_admin_user)
    return {"c": ERR_SUCCESS[0], "m": ERR_SUCCESS[1], "d": ret_admin_user_list}


def export_admin_user(user):
    rep_dict = list_admin_user(user)
    if rep_dict["c"] != ERR_SUCCESS[0]:
        return rep_dict
    admin_user_list = rep_dict["d"]

    # write to tmp file
    property_display_list = [u"管理员姓名", u"管理应用", u"管理内容", u"联系电话", u"最近登陆"]
    property_name_list = ["user_name", "service_name", "role_name", "mobile", "last_login"]
    file_path = gen_path()
    wb = openpyxl.Workbook(encoding='utf-8')
    ws = wb.active
    col = 1
    row = 1
    font = Font(bold=True)
    for property_display in property_display_list:
        cell = ws.cell(column=col, row=row, value=property_display)
        cell.font = font
        col += 1
    row += 1
    for admin_user in admin_user_list:
        col = 1
        for property_name in property_name_list:
            cell = ws.cell(column=col, row=row, value=admin_user[property_name])
            col += 1
        row += 1
    wb.save(filename=file_path)

    return {"c": ERR_SUCCESS[0], "m": ERR_SUCCESS[1], "d": [file_path]}


@transaction.atomic()
def update_admin_user(user, role_id_list, teacher_id):
    school_id = user.school_id
    role_id_list = json.loads(role_id_list)
    teacher_id = int(teacher_id)

    # 用户自己不能删除自己系统管理员的身份
    teacher_obj = Teacher.objects.filter(account_id=user.id, school_id=school_id, del_flag=FLAG_NO).first()
    if teacher_obj and teacher_id == teacher_obj.id:
        sys_admin_role = Role.objects.filter(service__code=SERVICE_CODE_USER_CENTER, code=ADMIN_USER_TYPE_SYSTEM_ADMIN, del_flag=FLAG_NO).first()
        sys_admin_role_id = str(sys_admin_role.id)
        if sys_admin_role_id not in role_id_list:
            role_id_list.append(sys_admin_role_id)

    # 取消该用户的所有角色
    UserRole.objects.filter(school_id=school_id, user_id=int(teacher_id), del_flag=FLAG_NO).\
        update(del_flag=FLAG_YES, update_time=datetime.datetime.now())

    # 获取该学校开通的所用服务的角色
    service_id_list = SchoolService.objects.filter(school_id=school_id, del_flag=NO).values_list("service_id", flat=True)
    available_role_id_list = Role.objects.filter(service_id__in=list(service_id_list), id__in=role_id_list, del_flag=NO).\
        order_by("service_id").values_list('id', flat=True)
    for role_id in available_role_id_list:
        user_role = UserRole.objects.filter(school_id=school_id, user_id=teacher_id, role_id=role_id).first()
        if user_role:
            user_role.del_flag = FLAG_NO
            user_role.save()
        else:
            UserRole.objects.create(school_id=school_id, user_id=teacher_id, role_id=role_id)
    return {"c": ERR_SUCCESS[0], "m": ERR_SUCCESS[1], "d": []}


@transaction.atomic()
def delete_admin_user(user, user_role_id_list):
    school_id = user.school_id
    user_role_id_list = json.loads(user_role_id_list)

    total_num = len(user_role_id_list)
    real_num = 0
    for user_role_id in user_role_id_list:
        user_role_obj = UserRole.objects.filter(school_id=school_id, id=user_role_id, del_flag=FLAG_NO).first()
        # 用户自己不能删除自己系统管理员的身份
        if user_role_obj.user.account_id == user.id \
                and user_role_obj.role.service.code == SERVICE_CODE_USER_CENTER \
                and str(user_role_obj.role.code) == ADMIN_USER_TYPE_SYSTEM_ADMIN\
                and str(user_role_obj.role_id) in user.role:
            if total_num == 1:
                return {"c": ERR_SYSTEM_ADMIN_NOT_DELETE_SELF_ERROR[0], "m": ERR_SYSTEM_ADMIN_NOT_DELETE_SELF_ERROR[1], "d": []}
            else:
                continue
        if user_role_obj:
            user_role_obj.del_flag = FLAG_YES
            user_role_obj.save()
            real_num += 1

    if real_num == total_num:
        return {"c": ERR_SUCCESS[0], "m": ERR_SUCCESS[1], "d": []}
    elif total_num == 1:
        return {"c": ERR_DELETE_ERROR[0], "m": ERR_DELETE_ERROR[1], "d": []}
    else:
        msg = u"%d条记录删除成功，%d条记录删除失败" % (real_num, total_num-real_num)
        return {"c": ERR_DELETE_PART[0], "m": msg, "d": []}


def list_role_user(user, service_name=""):
    school_id = user.school_id

    # 获取该学校开通的所用服务的角色
    service_id_list = SchoolService.objects.filter(school_id=school_id, del_flag=NO).values_list("service_id", flat=True)
    role_list = Role.objects.filter(service_id__in=list(service_id_list), del_flag=NO).all()
    if len(service_name) > 0:
        role_list = role_list.filter(service__name__contains=service_name, del_flag=NO)
    role_list = role_list.values('service__name', 'id', 'name')

    # 获取这些角色对应的用户
    role_id_list = map(lambda x: x['id'], role_list)
    admin_list = UserRole.objects.filter(school_id=school_id, role_id__in=role_id_list, del_flag=NO).\
        values('role_id', 'user__id', 'user__full_name')

    admin_dict = {}
    for admin in admin_list:
        account_info = dict(id=admin['user__id'],
                            full_name=admin['user__full_name'])
        if admin['role_id'] in admin_dict.keys():
            admin_dict[admin['role_id']].append(account_info)
        else:
            admin_dict[admin['role_id']] = [account_info]

    ret_role_list = []
    for role in role_list:
        ret_role = dict(service_name=role['service__name'], role_id=str(role['id']),
                        role_name=role['name'],
                        user_list=admin_dict.get(role['id'], []))
        ret_role_list.append(ret_role)
    dict_resp = {"c": ERR_SUCCESS[0], "m": ERR_SUCCESS[1], "d": ret_role_list}
    return dict_resp


def update_role_user(user, role_id, user_list):
    school_id = user.school_id
    role_id = int(role_id)

    # 检查学校是否开通该服务
    service_id_list = SchoolService.objects.filter(school_id=school_id, del_flag=NO).values_list("service_id", flat=True)
    role_id_list = Role.objects.filter(service_id__in=list(service_id_list), del_flag=NO).values_list("id", flat=True)
    if role_id not in role_id_list:
        return {"c": ERR_USER_AUTH[0], "m": ERR_USER_AUTH[1], "d": []}

    # 获取新增和删除的role user
    dst_user_id_list = map(lambda x: int(x), user_list)
    src_user_id_list = UserRole.objects.filter(school_id=school_id, role_id=role_id, del_flag=NO).values_list("user_id", flat=True)
    should_add = list(set(dst_user_id_list) - set(src_user_id_list))
    should_delete = list(set(src_user_id_list) - set(dst_user_id_list))

    # 用户自己不能删除自己系统管理员的身份
    role = Role.objects.filter(id=role_id).values("code", "service__code").first()
    if not role:
        return {"c": ERR_REQUEST_PARAMETER_ERROR[0], "m": ERR_REQUEST_PARAMETER_ERROR[1], "d": []}
    if role["service__code"] == SERVICE_CODE_USER_CENTER and str(role["code"]) == ADMIN_USER_TYPE_SYSTEM_ADMIN \
            and Teacher.objects.filter(account_id=user.id, id__in=should_delete).exists():
        return {"c": ERR_SYSTEM_ADMIN_NOT_DELETE_SELF_ERROR[0], "m": ERR_SYSTEM_ADMIN_NOT_DELETE_SELF_ERROR[1], "d": []}

    UserRole.objects.filter(role_id=role_id, user_id__in=should_delete, del_flag=NO).\
        update(del_flag=YES, update_time=datetime.datetime.now())
    new_user_role_list = []
    for user_id in should_add:
        new_user_role_list.append(UserRole(school_id=school_id, role_id=role_id, user_id=user_id))
    UserRole.objects.bulk_create(new_user_role_list)

    return {"c": ERR_SUCCESS[0], "m": ERR_SUCCESS[1], "d": []}


def get_service_user_role(service_code, user_id, school_id):
    role_code_list = UserRole.objects.filter(user_id=user_id, school_id=school_id, role__service__code=service_code, del_flag=NO)\
        .values_list("role__code", flat=True)
    # ret_role_code_list = map(lambda x: int(x), list(role_code_list))
    return list(role_code_list)


def init_service():
    # 检查数据是否已经初始化
    try:
        if Service.objects.filter(del_flag=NO).exists():
            return
    except ProgrammingError:  # syncdb时会抛出该异常
        return
    file_name = os.path.join(settings.BASE_DIR, "user_center", "settings", "service.json")
    try:
        f = file(file_name)
        service_info = json.load(f)

        # 初始化服务
        service_list = service_info["service"]
        service_code_list = list(Service.objects.filter(del_flag=NO).values_list("code", flat=True))
        for service in service_list:
            if service["code"] not in service_code_list:
                Service.objects.create(code=service["code"], name=service["name"], type=service["type"],
                                       intranet_url=service["intranet_url"], internet_url=service["internet_url"],
                                       access_key=service["access_key"], secret_key=service["secret_key"],
                                       comments=service["comments"], access_mask=service["access_mask"],
                                       is_cls_adviser_as_mgr=service["is_cls_adviser_as_mgr"],
                                       classify=service["classify"])

        # 初始化角色
        role_list = service_info["role"]
        db_role_list = Role.objects.filter(del_flag=NO).values("service__code", "code")
        for role in role_list:
            has_role = False
            for db_role in db_role_list:
                if db_role["service__code"] == role["service_code"] and db_role["code"] == role["role_code"]:
                    has_role = True
            if not has_role:
                service = Service.objects.filter(code=role["service_code"]).first()
                if service:
                    Role.objects.create(service=service, code=role["role_code"], name=role["name"])

    except Exception as ex:
        logger.error(u"无法读取服务角色配置文件")
        sErrInfo = traceback.format_exc()
        logger.error(sErrInfo)


def get_user_center_intra_url():
    user_center_url = Service.objects.filter(code=settings.SERVICE_USER_CENTER, del_flag=NO).first()
    if user_center_url:
        return get_domain_name(user_center_url.internet_url)
    else:
        return ""


def list_service_apps(user):
    school_id = user.school_id
    user_type = user.type
    user_center_url = get_user_center_intra_url()

    # 过滤学校开通的服务
    service_id_list = SchoolService.objects.filter(school_id=school_id, del_flag=NO).values_list("service_id")
    service_list = Service.objects.filter(id__in=service_id_list, del_flag=NO).\
        values("id", "code", "name", "type", "internet_url", "classify", "access_mask",
               "is_cls_adviser_as_mgr", "comments", "image__url")

    # 准备数据
    is_cls_adviser = False  # 是否是班级管理员
    mgr_service_id_list = []  # 管理的服务ID列表
    teacher_obj = Teacher.objects.filter(school_id=school_id, account_id=user.id, is_in=YES, del_flag=NO).first()
    if user_type == USER_TYPE_TEACHER and teacher_obj:
        if teacher_obj.cls:
            is_cls_adviser = True
        mgr_service_id_list = UserRole.objects.filter(school_id=school_id, user_id=teacher_obj.id, del_flag=NO).\
            values_list("role__service_id", flat=True)

    # 用户类型过滤服务
    available_service_list = []
    for service in service_list:
        # 若为管理员忽略user_type
        if service["id"] in mgr_service_id_list:
            available_service_list.append(service)
        elif is_cls_adviser and service["is_cls_adviser_as_mgr"]:
            available_service_list.append(service)
        elif user_type & service["access_mask"]:
            available_service_list.append(service)
        # elif service["access_mask"] == SERVICE_ACCESS_MASK_ONLY_MANAGER:


    # 转换输出格式
    category_service_dict = {}
    for service in available_service_list:
        image_url = ""
        if service["image__url"]:
            image_url = get_image_url(service["image__url"], user_center_url)
        category_name = SERVICE_CLASSIFY_DICT[service["classify"]]
        if category_name not in category_service_dict.keys():
            category_service_dict[category_name] = []
        if service["image__url"]:
            service["image_url"] = get_image_url(service["image__url"])
        category_service_dict[category_name].append({"id": str(service["id"]), "code": service["code"],
                                                     "name": service["name"], "url": service["internet_url"],
                                                     "type": str(service["type"]), "comments": service["comments"],
                                                     "image_url": image_url})
    ret_category_service_list = []
    for classify_id, classify_name in SERVICE_CLASSIFY_CHOICE:
        if category_service_dict.has_key(classify_name):
            ret_category_service = {"category": classify_name, "data": category_service_dict[classify_name]}
            ret_category_service_list.append(ret_category_service)
    # for key, value in category_service_dict.items():
    #     ret_category_service = {"category": key, "data": value}
    #     ret_category_service_list.append(ret_category_service)

    return {"c": ERR_SUCCESS[0], "m": ERR_SUCCESS[1], "d": ret_category_service_list}


