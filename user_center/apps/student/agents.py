# -*- coding=utf-8 -*-

from django.db.models import Q
from django.db import transaction

from models import *
from user_center.utils.constant import *
from user_center.utils.err_code import *
from user_center.utils.public_fun import *
from user_center.utils.request_auth import *
from user_center.apps.account.agents import *
from user_center.apps.school.models import *
from user_center.utils.file_fun import *
import openpyxl
from openpyxl.styles import Font


def graduate_student(school_id):
    Student.objects.filter(school_id=school_id, cls__graduate_status=YES, is_in=YES, del_flag=NO).\
        update(is_in=NO, kind=STUDENT_TYPE_GRADUATE, out_date=datetime.datetime.now(), update_time=datetime.datetime.now())


def undo_graduate_student(school_id):
    Student.objects.filter(school_id=school_id, kind=STUDENT_TYPE_GRADUATE, cls__graduate_status=NO, del_flag=NO).\
        update(is_in=YES, kind=STUDENT_TYPE_NORMAL, out_date=None, update_time=datetime.datetime.now())


def list_student(user=None, full_name="", code="", name_or_code="", is_in="", kind="", is_available="", grade_name="",
                 class_id="", enrollment_year="", graduated_year="", verbose="", school_id=""):
    if school_id:
        school_id = int(school_id)
    else:
        school_id = user.school_id

    # 检查用户角色班主任只能管理本班级学生
    if user and ADMIN_USER_TYPE_CLASS_ADMIN in user.role:
        filter_class_id = -1
        if class_id:
            filter_class_id = int(class_id)
        dict_resp = check_cls_permission(user, filter_class_id)
        if dict_resp["c"] != ERR_SUCCESS[0]:
            return dict_resp
        else:
            class_id = dict_resp["d"][0]

    student_list = Student.objects.filter(school_id=school_id, del_flag=NO).all()
    if verbose:
        verbose = int(verbose)
        if verbose == 0:  # 在读学生
            student_list = student_list.filter(cls__graduate_status=NO)
        elif verbose == 1:  # 已毕业学生
            student_list = student_list.filter(cls__graduate_status=YES)
        elif verbose == 2:  # 未分配班级学生
            student_list = student_list.filter(cls__isnull=True)
        else:  # 所有学生
            pass
    else:
        student_list = student_list.filter(cls__graduate_status=NO)
    if not verbose or verbose == "0":  # 默认不显示已经毕业学生信息
        # student_list = student_list.filter(out_date__isnull=True)
        student_list = student_list.filter(cls__graduate_status=NO)

    if full_name:
        student_list = student_list.filter(full_name__contains=full_name)
    if code:
        student_list = student_list.filter(Q(account__code__contains=code) | Q(account__tmp_code__contains=code))
    if name_or_code:
        student_list = student_list.filter(Q(account__code__contains=name_or_code) |
                                           Q(account__tmp_code__contains=name_or_code) |
                                           Q(full_name__contains=name_or_code))
    if is_in:
        is_in = str_to_int(is_in)
        if is_in > 0:
            student_list = student_list.filter(is_in=is_in)
    if kind:
        student_list = student_list.filter(kind=kind)
    if is_available:
        is_available = str_to_int(is_available)
        if is_available > 0:
            student_list = student_list.filter(is_available=is_available)
    if grade_name:
        student_list = student_list.filter(cls__grade_name=grade_name)
    if class_id:
        class_id = str_to_int(class_id)
        if class_id > 0:
            student_list = student_list.filter(cls_id=class_id)
    if enrollment_year:
        student_list = student_list.filter(cls__enrollment_year=enrollment_year)
    if graduated_year:
        student_list = student_list.filter(cls__graduated_year=graduated_year)

    student_list = student_list.values('id', 'account_id', 'account__username', 'full_name', 'id_card',
                                       'account__code', 'sex', 'account__mobile', 'email',
                                       'cls__class_name', 'cls__grade_name', 'entry_date', 'out_date',
                                       'is_in', 'kind', 'comments', "image__url", 'account__last_login')
    ret_student_list = []
    for student in student_list:
        ret_student = dict(id=str(student['id']), account_id=str(student['account_id']), username=student['account__username'],
                           full_name=student['full_name'], id_card=student['id_card'],
                           code=student['account__code'], sex=student['sex'],
                           mobile=student['account__mobile'], email=student['email'],
                           class_name=student['cls__class_name'] or "",
                           grade_name=student['cls__grade_name'] or "",
                           entry_date=datetime_to_str(student['entry_date']),
                           out_date=datetime_to_str(student['out_date']),
                           is_in=str(student['is_in']), kind=student['kind'],
                           comments=student['comments'], image_url="",
                           last_login=datetime_to_str(student['account__last_login'], DATE_FORMAT_TIME))
        if student["image__url"]:
            ret_student["image_url"] = get_image_url(student["image__url"])
        ret_student_list.append(ret_student)
    dict_resp = {"c": ERR_SUCCESS[0], "m": ERR_SUCCESS[1], "d": ret_student_list}
    return dict_resp


def detail_student(user, student_id, account_id):
    school_id = user.school_id

    student_obj = Student.objects.filter(id=int(student_id), school_id=school_id, del_flag=NO)
    student_info = student_obj.values("id", "account_id", "account__username", "full_name", "id_card",
                                      "account__code", "account__tmp_code", "sex", "account__mobile",
                                      "email", "address", "cls_id", "cls__class_name", "cls__grade_name", "entry_date",
                                      "out_date", "is_in", "kind", "comments", "image__url", "birthday",
                                      "native_place").first()

    # 检查自己修改自己的信息
    if account_id and int(account_id) != student_info["account_id"]:
        return {"c": ERR_USER_AUTH[0], "m": ERR_USER_AUTH[1], "d": []}

    # 检查用户角色班主任只能管理本班级学生
    if ADMIN_USER_TYPE_CLASS_ADMIN in user.role:
        dict_resp = check_cls_permission(user, student_info["cls_id"])
        if dict_resp["c"] != ERR_SUCCESS[0]:
            return dict_resp

    ret_student_info = dict(id=str(student_info["id"]),
                            account_id=str(student_info['account_id']),
                            username=student_info["account__username"],
                            full_name=student_info["full_name"],
                            id_card=student_info["id_card"],
                            code=student_info["account__code"],
                            tmp_code=student_info["account__tmp_code"],
                            sex=student_info["sex"],
                            mobile=student_info["account__mobile"],
                            email=student_info["email"],
                            address=student_info["address"],
                            class_id=student_info["cls_id"] or "",
                            class_name=student_info["cls__class_name"] or "",
                            grade_name=student_info["cls__grade_name"] or "",
                            entry_date=datetime_to_str(student_info["entry_date"]),
                            out_date=datetime_to_str(student_info["out_date"]),
                            is_in=str(student_info["is_in"]),
                            kind=student_info["kind"],
                            comments=student_info["comments"],
                            birthday=datetime_to_str(student_info["birthday"]),
                            native_place=student_info["native_place"],
                            image_url=""
                            )
    if student_info["image__url"]:
        ret_student_info["image_url"] = get_image_url(student_info["image__url"])
    return {"c": ERR_SUCCESS[0], "m": ERR_SUCCESS[1], "d": [ret_student_info]}


@transaction.atomic
def add_student(user, student_info, is_override=False):
    school_id = user.school_id

    username = student_info.get("username", "")
    password = student_info.get("password", "")
    code = student_info.get("code", "")
    # tmp_code = student_info.get("tmp_code", "")
    mobile = student_info.get("mobile", "")

    full_name = clean_string(student_info.get("full_name", ""))
    id_card = student_info.get("id_card", "")
    sex = student_info.get("sex", "")
    email = student_info.get("email", "")
    address = student_info.get("address", "")
    kind = student_info.get("kind", STUDENT_TYPE_NORMAL)
    is_in = student_info.get("is_in", "")
    is_available = student_info.get("is_available", "")
    entry_date = student_info.get("entry_date", "")
    out_date = student_info.get("out_date", "")
    comments = student_info.get("comments", "")
    birthday = student_info.get("birthday", "")
    native_place = student_info.get("native_place", "")
    image_url = student_info.get("image_url", "")
    image_obj = None
    if image_url:
        image_url = get_file_name_from_url(image_url)
        image_obj = Image.objects.filter(url__contains=image_url, del_flag=NO).first()
        if not image_obj:
            return {"c": ERR_USER_INFO_IMAGE_NOT_EXIST[0], "m": ERR_USER_INFO_IMAGE_NOT_EXIST[1], "d": []}

    # 在读状态、学生类型、出校时间校验和更新
    if is_in and int(is_in) == NO:
        if kind not in OUT_SCHOOL_STUDENT_TYPE_LIST:
            return {"c": ERR_STUDENT_KIND_NOT_MATCH_ERROR[0], "m": ERR_STUDENT_KIND_NOT_MATCH_ERROR[1], "d": []}
        out_date = datetime_to_str(datetime.datetime.now())
    elif is_in and int(is_in) == YES:
        if kind and kind not in IN_SCHOOL_STUDENT_TYPE_LIST:
            return {"c": ERR_STUDENT_KIND_NOT_MATCH_ERROR[0], "m": ERR_STUDENT_KIND_NOT_MATCH_ERROR[1], "d": []}
    else:
        is_in = YES
        kind = STUDENT_TYPE_NORMAL

    # 检查必填项
    if not code or not full_name:
        logger.error(" parent info is not integrity")
        return ERR_USER_INFO_INCOMPLETE

    # 检查班级是否正确并获取class_id
    class_name = student_info.get("class_name", "")
    class_id = student_info.get("class_id", "")
    cls = None
    if class_id:
        cls = Class.objects.filter(school_id=school_id, id=class_id, del_flag=NO).first()
        if not cls:
            return {"c": ERR_CLASS_NOT_EXIST[0], "m": ERR_CLASS_NOT_EXIST[1], "d": []}
    elif class_name:
        cls = Class.objects.filter(school_id=school_id, class_name=class_name, graduate_status=NO, del_flag=NO).first()
        if not cls:
            return {"c": ERR_CLASS_NOT_EXIST[0], "m": ERR_CLASS_NOT_EXIST[1], "d": []}

    # 检查用户角色班主任只能管理本班级学生
    if ADMIN_USER_TYPE_CLASS_ADMIN in user.role:
        cls_id = None
        if cls:
            cls_id = cls.id
        dict_resp = check_cls_permission(user, cls_id)
        if dict_resp["c"] != ERR_SUCCESS[0]:
            return dict_resp

    # 检查帐号在同一个学校的同一个角色是否已经存在
    if Student.objects.filter(school_id=school_id, account__code=code, del_flag=NO).exists():
        return {"c": ERR_USER_ALREADY_EXIST[0], "m": ERR_USER_ALREADY_EXIST[1], "d": []}

    # 创建帐号或获取已存在的帐号
    err_code = add_account(code=code, password=password, type=USER_TYPE_STUDENT, school_id=school_id)
    if err_code[0] != ERR_SUCCESS[0]:
        return {"c": err_code[0], "m": err_code[1], "d": []}
    account_id = err_code[1]

    # 检查手机号对应的帐号和学号对应的帐号是否为一个人
    if mobile:
        account = Account.objects.filter(mobile=mobile, del_flag=NO).first()
        if account and account.id != account_id:
            return {"c": ERR_USER_MOBILE_CONFLICT_ERROR[0], "m": ERR_USER_MOBILE_CONFLICT_ERROR[1], "d": []}
        else:
            account = Account.objects.get(id=account_id)
            update_account(user=user, account=account, mobile=mobile)

    # 检查身份证号是否已经占用
    if id_card and check_id_code_have_occupied(id_card, account_id):
        raise Exception(ERR_USER_ID_CARD_CONFLICT_ERROR[1])

    # 添加学生到Student表
    student = Student.objects.create(school_id=school_id, account_id=account_id)
    if full_name: student.full_name = full_name
    if id_card: student.id_card = id_card
    if sex: student.sex = sex
    if email: student.email = email
    if address: student.address = address
    if kind: student.kind = kind
    if cls: student.cls = cls
    if is_in: student.is_in = int(is_in)
    if is_available: student.is_available = int(is_available)
    if entry_date: student.entry_date = str_to_datetime(entry_date)
    if out_date: student.out_date = str_to_datetime(out_date)
    if comments: student.comments = comments
    if birthday: student.birthday = str_to_datetime(birthday)
    if native_place: student.native_place = native_place
    if image_obj: student.image = image_obj
    student.save()
    # 更新班级中的学生计数
    if cls:
        refresh_student_amount(school_id, [cls])

    return {"c": ERR_SUCCESS[0], "m": ERR_SUCCESS[1], "d": []}


@transaction.atomic
def delete_student(user, student_id_list):
    school_id = user.school_id

    # id_list and account_id_list should be delete
    id_list = []
    account_id_list = []
    student_id_list = map(lambda x: int(x), student_id_list)
    student_list = Student.objects.filter(id__in=student_id_list, school_id=school_id,
                                          del_flag=NO).values("id", "account_id", "cls_id")
    for student in student_list:
        # 检查用户角色班主任只能管理本班级学生
        if ADMIN_USER_TYPE_CLASS_ADMIN in user.role:
            dict_resp = check_cls_permission(user, student["cls_id"])
            if dict_resp["c"] != ERR_SUCCESS[0]:
                continue
        id_list.append(student["id"])
        account_id_list.append(student["account_id"])

    # delete student
    Student.objects.filter(id__in=list(id_list), del_flag=NO).update(del_flag=YES, update_time=datetime.datetime.now())
    ParentStudent.objects.filter(student_id__in=list(id_list), del_flag=NO).update(del_flag=YES, update_time=datetime.datetime.now())

    # delete Account if necessary
    for account_id in account_id_list:
        check_and_delete_account(account_id)
    refresh_student_amount(school_id=school_id)
    # check if delete all
    err_num = len(student_id_list) - len(id_list)
    if err_num > 0:
        msg = u"%d条记录删除失败，可能是权限不够" % err_num
        dict_resp = {"c": -1, "m": msg, "d": []}
        return dict_resp
    return {"c": ERR_SUCCESS[0], "m": ERR_SUCCESS[1], "d": []}


@transaction.atomic
def update_student_class(user, student_id_list, class_id):
    school_id = user.school_id

    # id_list should be delete
    student_id_list = map(lambda x: int(x), student_id_list)
    student_list = Student.objects.filter(id__in=student_id_list, school_id=school_id, del_flag=NO)

    # 检查班级是否正确并获取class_id
    if class_id:
        class_id = int(class_id)
        cls = Class.objects.filter(school_id=school_id, id=class_id, del_flag=NO).first()
        if not cls:
            return {"c": ERR_CLASS_NOT_EXIST[0], "m": ERR_CLASS_NOT_EXIST[1], "d": []}
    else:
        cls = None

    total_num = len(student_id_list)
    real_num = 0
    for student_obj in student_list:
        student_obj.cls = cls
        student_obj.save()
        real_num += 1
    refresh_student_amount(school_id=school_id)

    # return msg
    if real_num == total_num:
        return {"c": ERR_SUCCESS[0], "m": ERR_SUCCESS[1], "d": []}
    elif total_num == 1:
        return {"c": ERR_OP_ERROR[0], "m": ERR_OP_ERROR[1], "d": []}
    else:
        msg = u"%d条记录操作失败，%d条记录操作成功" % (total_num-real_num, real_num)
        return {"c": ERR_OP_PART_ERROR[0], "m": msg, "d": []}


@transaction.atomic
def update_student(user, student_info, account_id):
    school_id = user.school_id

    student_id = student_info.get("id", "")
    if not student_id:
        return {"c": ERR_USER_NOT_EXIST[0], "m": ERR_USER_NOT_EXIST[1], "d": []}
    student = Student.objects.filter(id=student_id, school_id=school_id, del_flag=NO).first()
    if not student:
        return {"c": ERR_USER_NOT_EXIST[0], "m": ERR_USER_NOT_EXIST[1], "d": []}

    # 检查自己修改自己的信息
    if account_id and int(account_id) != student.account_id:
        return {"c": ERR_USER_AUTH[0], "m": ERR_USER_AUTH[1], "d": []}

    username = student_info.get("username")
    code = student_info.get("code")
    mobile = student_info.get("mobile")
    full_name = clean_string(student_info.get("full_name"))
    sex = student_info.get("sex")
    email = student_info.get("email")
    id_card = student_info.get("id_card")
    address = student_info.get("address")
    image_url = student_info.get("image_url")
    class_id = student_info.get("class_id")
    kind = student_info.get("kind")
    is_in = student_info.get("is_in")
    is_available = student_info.get("is_available")
    entry_date = student_info.get("entry_date")
    out_date = student_info.get("out_date")
    comments = student_info.get("comments")
    birthday = student_info.get("birthday", "")
    native_place = student_info.get("native_place", "")

    # 检查身份证号是否已经占用
    if id_card and check_id_code_have_occupied(id_card, student.account_id):
        raise Exception(ERR_USER_ID_CARD_CONFLICT_ERROR[1])

    image_obj = None
    if image_url:
        image_url = get_file_name_from_url(image_url)
        image_obj = Image.objects.filter(url__contains=image_url, del_flag=NO).first()
        if not image_obj:
            return {"c": ERR_USER_INFO_IMAGE_NOT_EXIST[0], "m": ERR_USER_INFO_IMAGE_NOT_EXIST[1], "d": []}
    if class_id:
        if not Class.objects.filter(id=class_id, del_flag=NO).exists():
            return {"c": ERR_CLASS_NOT_EXIST[0], "m": ERR_CLASS_NOT_EXIST[1], "d": []}

    # 检查用户角色班主任只能管理本班级学生
    if ADMIN_USER_TYPE_CLASS_ADMIN in user.role:
        real_class_id = student.cls_id
        if class_id:
            real_class_id = int(class_id)
        dict_resp = check_cls_permission(user, real_class_id)
        if dict_resp["c"] != ERR_SUCCESS[0]:
            return dict_resp

    # 在读状态、学生类型、出校时间校验和更新
    if is_in and int(is_in) == NO:
        if kind not in OUT_SCHOOL_STUDENT_TYPE_LIST:
            return {"c": ERR_STUDENT_KIND_NOT_MATCH_ERROR[0], "m": ERR_STUDENT_KIND_NOT_MATCH_ERROR[1], "d": []}
        out_date = datetime_to_str(datetime.datetime.now())
    elif is_in and int(is_in) == YES:
        if kind and kind not in IN_SCHOOL_STUDENT_TYPE_LIST:
            return {"c": ERR_STUDENT_KIND_NOT_MATCH_ERROR[0], "m": ERR_STUDENT_KIND_NOT_MATCH_ERROR[1], "d": []}

    account = student.account
    err_code = update_account(user=user, account=account, username=username, code=code, mobile=mobile)
    if err_code != ERR_SUCCESS:
        return {"c": err_code[0], "m": err_code[1], "d": []}

    # if code and code != account.code:
    #     err_code = add_account(code=code, mobile=mobile, type=USER_TYPE_STUDENT, school_id=school_id)
    #     if err_code[0] != ERR_SUCCESS[0]:
    #         return {"c": err_code[0], "m": err_code[1], "d": []}
    #     account_id = err_code[1]
    #     check_and_delete_account(account.id)
    src_cls_id = student.cls_id

    if full_name: student.full_name = full_name
    if id_card: student.id_card = id_card
    elif id_card == "": student.id_card = id_card
    if sex: student.sex = sex
    if email: student.email = email
    elif email == "": student.email = email
    if address: student.address = address
    elif address == "": student.address = address
    if image_obj: student.image = image_obj
    if kind: student.kind = kind
    elif kind == "": student.kind = kind
    if class_id: student.cls_id = int(class_id)
    if is_in: student.is_in = int(is_in)
    if is_available: student.is_available = int(is_available)
    if entry_date: student.entry_date = str_to_datetime(entry_date)
    if out_date: student.out_date = str_to_datetime(out_date)
    if comments: student.comments = comments
    elif comments == "": student.comments = comments
    if image_obj: student.image = image_obj
    if birthday: student.birthday = str_to_datetime(birthday)
    if native_place: student.native_place = native_place
    student.save()

    if is_in or (class_id and src_cls_id != int(class_id)):
        dst_cls_obj = Class.objects.get(id=int(class_id))
        src_cls_obj = Class.objects.get(id=src_cls_id)
        refresh_student_amount(school_id, [src_cls_obj, dst_cls_obj])
    # 更新学生表信息
    return {"c": ERR_SUCCESS[0], "m": ERR_SUCCESS[1], "d": []}


def import_student(user, file_obj, is_override=False):
    school_id = user.school_id
    ret_data = []
    skip_id_list = []
    success_num = 0

    if not file_obj.name.endswith('xlsx'):
        return {"c": ERR_FILE_FORMAT_NOT_SUPPORTED[0], "m": ERR_FILE_FORMAT_NOT_SUPPORTED[1],  "d": []}

    # write file data to tmp path
    file_path = gen_path()
    with open(file_path, 'wb+') as destination:
        for chunk in file_obj.chunks():
            destination.write(chunk)

    wb = openpyxl.load_workbook(file_path)
    ws = wb.get_active_sheet()

    property_display_list = [u"姓名*", u"性别*", u"学籍号*", u"班级名称", u"手机号", u"身份证号"]
    property_name_list = ["full_name", "sex", "code", "class_name", "mobile", "id_card"]
    property_len = len(property_name_list)

    # check the first row
    for i in range(property_len):
        if property_display_list[i] != ws.cell(column=i+1, row=1).value:
            return {"c": ERR_FILE_TEMPLATE_ERROR[0], "m": ERR_FILE_TEMPLATE_ERROR[1],  "d": []}
    row_num = 0

    for row in ws.rows:
        row_num += 1
        if row_num == 1:
            # 第一行为标题栏
            continue
        col_num = 0
        student_info = {}
        for cell in row:
            if col_num >= property_len:
                break
            property_name = property_name_list[col_num]
            value = ""
            if isinstance(cell.value, (int, long)):
                value = str(cell.value)
            elif isinstance(cell.value, basestring):
                value = cell.value.strip()
            student_info[property_name] = value
            col_num += 1
        # ignore blank line
        if not student_info["full_name"]:
            continue
        # check not null field
        if not student_info["sex"] or not student_info["code"]:
            err_msg = u"第%d行: 用户[姓名/性别/学籍号]为空" % row_num
            ret_data.append(err_msg)
            continue
        try:
            err_code = add_student(user, student_info, is_override)
        except Exception as ex:
            err_code = {"c": -1, "m": ex.message}
        if err_code['c'] == ERR_USER_ALREADY_EXIST[0]:
            skip_id_list.append(row_num)
        elif err_code['c'] != ERR_SUCCESS[0]:
            err_msg = u"第%d行[%s]: " % (row_num, student_info["full_name"])
            err_msg += err_code['m']
            ret_data.append(err_msg)
        else:
            success_num += 1
    # return message
    if ret_data:
        ret_msg = u"%d条记录新增成功，%d条记录重复跳过，%d条记录新增失败" % (success_num, len(skip_id_list), len(ret_data))
        return {"c": ERR_IMPORT_DATA_ERROR[0], "m": ret_msg, "d": ret_data}
    else:
        ret_msg = u"%d条记录新增成功，%d条记录重复跳过" % (success_num, len(skip_id_list))
        return {"c": ERR_SUCCESS[0], "m": ret_msg, "d": []}


def export_student(user, full_name="", code="", name_or_code="", is_in="", kind="", is_available="",
                   grade_name="", class_id="",  enrollment_year="", verbose=""):
    ret_resp = list_student(user=user, full_name=full_name, code=code, name_or_code=name_or_code,
                            is_in=is_in, kind=kind, is_available=is_available, grade_name=grade_name,
                            class_id=class_id, enrollment_year=enrollment_year, verbose=verbose)
    if ret_resp["c"] != ERR_SUCCESS[0]:
        return ret_resp
    student_list = ret_resp["d"]

    # write to tmp file
    property_display_list = [u"姓名", u"性别", u"学籍号", u"年级名称", u"班级名称", u"联系方式", u"身份证号", u"入校时间"]
    property_name_list = ["full_name", "sex", "code", "grade_name", "class_name", "mobile", "id_card", "entry_date"]
    file_path = gen_path()
    wb = openpyxl.Workbook(encoding='utf-8')
    ws = wb.active
    col = 1
    row = 1
    font = Font(bold=True)
    for property_display in property_display_list:
        cell = ws.cell(column=col, row=row, value=property_display)
        cell.font = font
        col += 1
    row += 1
    for student in student_list:
        col = 1
        for property_name in property_name_list:
            cell = ws.cell(column=col, row=row, value=student[property_name])
            col += 1
        row += 1

    reset_excel_column_with(ws)
    wb.save(filename=file_path)

    return {"c": ERR_SUCCESS[0], "m": ERR_SUCCESS[1], "d": [file_path]}


# 更新班级中的学生总数
def refresh_student_amount(school_id, cls_obj_list=None):
    if not cls_obj_list:
        cls_obj_list = Class.objects.filter(school_id=school_id, del_flag=FLAG_NO)
    for cls_obj in cls_obj_list:
        student_amount = Student.objects.filter(school_id=school_id, cls=cls_obj, is_in=FLAG_YES, del_flag=FLAG_NO).count()
        cls_obj.student_amount = student_amount
        cls_obj.save()


def add_student_class_application(user, class_code, comments):
    school_id = user.school_id
    class_id = int(class_code)
    class_obj = Class.objects.filter(id=class_id, school_id=school_id, del_flag=FLAG_NO).first()
    if not class_obj:
        return {"c": ERR_CLASS_CODE_ERROR[0], "m": ERR_CLASS_CODE_ERROR[1], "d": []}
    student_obj = Student.objects.filter(school_id=school_id, account_id=user.id, del_flag=FLAG_NO).first()
    if not student_obj:
        return {"c": ERR_USER_NOT_EXIST[0], "m": ERR_USER_NOT_EXIST[1], "d": []}
    StudentClassApplication.objects.create(student=student_obj, cls=class_obj, status=APPLICATION_STATUS_NOT_PROCESS[0], comments=comments)
    return {"c": ERR_SUCCESS[0], "m": ERR_SUCCESS[1], "d": []}


def list_student_class_application(user, status, student_name):
    school_id = user.school_id

    teacher_obj = Teacher.objects.filter(school_id=school_id, account_id=user.id, del_flag=FLAG_NO).first()
    if not teacher_obj:
        return {"c": ERR_USER_NOT_EXIST[0], "m": ERR_USER_NOT_EXIST[1], "d": []}
    application_list = StudentClassApplication.objects.filter(cls=teacher_obj.cls, del_flag=FLAG_NO)
    if status:
        application_list = application_list.filter(status=int(status))
    if student_name:
        application_list = application_list.filter(student__full_name__contains=student_name)
    application_list = application_list.values("id", "student__account__code", "student__full_name", "comments",
                                               "status", "student__image__url")
    ret_application_list = []
    for application in application_list:
        ret_application = {"id": application["id"], "student_code": application["student__account__code"],
                           "student_full_name": application["student__full_name"], "comments": application["comments"],
                           "status": application["status"], 'image_url': get_image_url(application["student__image__url"])}

        ret_application_list.append(ret_application)
    return {"c": ERR_SUCCESS[0], "m": ERR_SUCCESS[1], "d": ret_application_list}


def delete_student_class_application(user, id_list):
    school_id = user.school_id
    id_list = map(lambda x: int(x), id_list)

    teacher_obj = Teacher.objects.filter(school_id=school_id, account_id=user.id, del_flag=FLAG_NO).first()
    if not teacher_obj:
        return {"c": ERR_USER_NOT_EXIST[0], "m": ERR_USER_NOT_EXIST[1], "d": []}

    application_list = StudentClassApplication.objects.filter(cls=teacher_obj.cls, id__in=id_list, del_flag=FLAG_NO).\
        exclude(status=APPLICATION_STATUS_NOT_PROCESS[0])
    total_num = len(id_list)
    real_num = 0
    for application in application_list:
        application.del_flag = FLAG_YES
        application.save()
        real_num += 1

    if total_num == real_num:
        return {"c": ERR_SUCCESS[0], "m": ERR_SUCCESS[1], "d": []}
    elif total_num == 1:
        return {"c": ERR_DELETE_ERROR[0], "m": ERR_DELETE_ERROR[1], "d": []}
    else:
        err_msg = u"%d删除成功，%d删除失败" % (real_num, total_num-real_num)
        return {"c": ERR_DELETE_PART[0], "m": err_msg, "d": []}


@transaction.atomic
def update_student_class_application(user, status, id_list):
    school_id = user.school_id
    status = int(status)
    id_list = map(lambda x: int(x), id_list)

    teacher_obj = Teacher.objects.filter(school_id=school_id, account_id=user.id, del_flag=FLAG_NO).first()
    if not teacher_obj:
        return {"c": ERR_USER_NOT_EXIST[0], "m": ERR_USER_NOT_EXIST[1], "d": []}
    if int(status) != APPLICATION_STATUS_APPROVED[0] and int(status) != APPLICATION_STATUS_REFUSED[0]:
        return {"c": ERR_OP_ERROR[0], "m": ERR_OP_ERROR[1], "d": []}

    application_list = StudentClassApplication.objects.filter(cls=teacher_obj.cls, id__in=id_list,
                                                              status=APPLICATION_STATUS_NOT_PROCESS[0], del_flag=FLAG_NO)
    total_num = len(id_list)
    real_num = 0
    for application in application_list:
        if status == APPLICATION_STATUS_APPROVED[0]:
            application.student.cls = application.cls
            application.student.save()
            refresh_student_amount(school_id, [application.cls])
        application.status = int(status)
        application.save()
        real_num += 1

    if total_num == real_num:
        return {"c": ERR_SUCCESS[0], "m": ERR_SUCCESS[1], "d": []}
    elif total_num == 1:
        return {"c": ERR_OP_ERROR[0], "m": ERR_OP_ERROR[1], "d": []}
    else:
        err_msg = u"%d操作成功，%d操作失败" % (real_num, total_num-real_num)
        return {"c": ERR_DELETE_PART[0], "m": err_msg, "d": []}


def detail_student_class_application(user):
    school_id = user.school_id

    student_obj = Student.objects.filter(school_id=school_id, account_id=user.id, del_flag=FLAG_NO).first()
    if not student_obj:
        return {"c": ERR_USER_NOT_EXIST[0], "m": ERR_USER_NOT_EXIST[1], "d": []}
    cls_info = {"class_id": "", "class_name": "", "class_master": "", "status": APPLICATION_STATUS_APPROVED[0]}
    if student_obj.cls:
        cls_obj = student_obj.cls
        cls_info["class_id"] = cls_obj.id
        cls_info["class_name"] = cls_obj.class_name
    else:
        application = StudentClassApplication.objects.filter(
            student=student_obj,
            status__in=[APPLICATION_STATUS_NOT_PROCESS[0], APPLICATION_STATUS_REFUSED[0]],
            del_flag=FLAG_NO).order_by("-id").first()
        if application:
            cls_info["class_id"] = application.cls.id
            cls_info["class_name"] = application.cls.class_name
            cls_info["status"] = application.status

    # 获取班主任信息
    if cls_info["class_id"]:
        teacher_obj = Teacher.objects.filter(school_id=school_id, cls_id=cls_info["class_id"], del_flag=FLAG_NO).first()
        if teacher_obj:
            cls_info["class_master"] = teacher_obj.full_name
    return {"c": ERR_SUCCESS[0], "m": ERR_SUCCESS[1], "d": [cls_info]}